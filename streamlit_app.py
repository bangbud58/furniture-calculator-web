"""
Furniture Price Generator - Streamlit Web Version
Bisa diakses dari browser (Windows/Android/iOS/Mac)
"""
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import io
from material_descriptions import MaterialDescriptions

# Page config
st.set_page_config(
    page_title="Furniture Price Generator",
    page_icon="🛋️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS - Red theme
st.markdown("""
<style>
    .main {
        background-color: #f5f6fa;
    }
    .stButton>button {
        background-color: #e74c3c;
        color: white;
        border-radius: 4px;
        padding: 0.5rem 1rem;
        font-size: 16px;
        border: none;
    }
    .stButton>button:hover {
        background-color: #ff6b6b;
    }
    h1 {
        color: #c0392b;
    }
    h2, h3 {
        color: #e74c3c;
    }
    .stExpander {
        border: 1px solid #dcdde1;
        border-radius: 4px;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'items_data' not in st.session_state:
    st.session_state.items_data = []

# Helper functions
def calculate_cabinet_price(panjang, tinggi, tebal, finishing):
    """Hitung harga kabinet dengan minimum dimensi 100cm"""
    panjang_calc = max(panjang, 100)
    tinggi_calc = max(tinggi, 100)
    
    area = (panjang_calc * tinggi_calc) / 10000
    
    thickness_multiplier = 1.0
    if tebal >= 3:
        thickness_multiplier = 1.5
    
    if finishing == 'HPL':
        price = 1800000
    elif finishing == 'Duco':
        price = 2200000
    else:  # Kombinasi
        price = 2000000
    
    unit_price = price * thickness_multiplier
    total = area * unit_price
    
    return total, area, unit_price

def calculate_surface_price(panjang, tinggi, material):
    """Hitung harga top table / backsplash"""
    area = (panjang * tinggi) / 10000
    
    material_prices = {
        'Solid Surface': 2500000,
        'Granit Alam': 1800000,
        'Marmer': 3500000,
        'Mirror Clear': 1200000,
        'Bronze Mirror': 1300000,
        'Keramik': 800000
    }
    
    unit_price = material_prices.get(material, 0)
    total = area * unit_price
    
    return total, area, unit_price

def calculate_wardrobe_price(panjang, tinggi, tebal, finishing):
    """Hitung harga wardrobe"""
    area = (panjang * tinggi) / 10000
    
    thickness_multiplier = 1.0
    if tebal >= 3:
        thickness_multiplier = 1.5
    
    if finishing == 'HPL':
        price = 1600000
    elif finishing == 'Duco':
        price = 2000000
    else:  # Kombinasi
        price = 1800000
    
    unit_price = price * thickness_multiplier
    total = area * unit_price
    
    return total, area, unit_price

def calculate_bed_price(panjang, lebar, tinggi_sandaran, finishing_sandaran, material_dudukan):
    """Hitung harga bed"""
    area_sandaran = (panjang * tinggi_sandaran) / 10000
    area_dudukan = (panjang * lebar) / 10000
    
    if finishing_sandaran == 'HPL':
        price_sandaran = 1500000
    elif finishing_sandaran == 'Duco':
        price_sandaran = 1900000
    else:  # Kombinasi
        price_sandaran = 1700000
    
    material_prices = {
        'Synthetic Leather': 1800000,
        'Fabric': 1200000
    }
    price_dudukan = material_prices.get(material_dudukan, 0)
    
    total_sandaran = area_sandaran * price_sandaran
    total_dudukan = area_dudukan * price_dudukan
    total = total_sandaran + total_dudukan
    
    return total, area_sandaran, area_dudukan, price_sandaran, price_dudukan

def calculate_backdrop_price(panjang, tinggi, finishing):
    """Hitung harga backdrop"""
    area = (panjang * tinggi) / 10000
    
    if finishing == 'HPL':
        price = 1400000
    elif finishing == 'Duco':
        price = 1800000
    else:  # Kombinasi
        price = 1600000
    
    total = area * price
    return total, area, price

def calculate_credenza_price(panjang, tinggi, tebal, finishing):
    """Hitung harga credenza"""
    area = (panjang * tinggi) / 10000
    
    thickness_multiplier = 1.0
    if tebal >= 3:
        thickness_multiplier = 1.5
    
    if finishing == 'HPL':
        price = 1700000
    elif finishing == 'Duco':
        price = 2100000
    else:  # Kombinasi
        price = 1900000
    
    unit_price = price * thickness_multiplier
    total = area * unit_price
    
    return total, area, unit_price

def calculate_multi_cabinet_price(panjang, tinggi, tebal, finishing):
    """Hitung harga multi cabinet"""
    area = (panjang * tinggi) / 10000
    
    thickness_multiplier = 1.0
    if tebal >= 3:
        thickness_multiplier = 1.5
    
    if finishing == 'HPL':
        price = 1600000
    elif finishing == 'Duco':
        price = 2000000
    else:  # Kombinasi
        price = 1800000
    
    unit_price = price * thickness_multiplier
    total = area * unit_price
    
    return total, area, unit_price

def export_to_excel():
    """Export items to Excel file with company header"""
    if not st.session_state.items_data:
        return None
    
    from openpyxl.styles import Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Harga"
    
    # ========== COMPANY HEADER ==========
    ws.merge_cells('A1:I1')
    ws.merge_cells('A2:I2')
    ws.merge_cells('A3:I3')
    ws.merge_cells('A4:I4')
    
    # Company name
    ws['A1'] = 'PT. ENIGMA PRISMA DELAPAN'
    ws['A1'].font = Font(bold=True, size=16, color='c0392b')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Address line 1
    ws['A2'] = 'Jl. Raya H. Abdullah No.56, Pakulonan Barat, Kelapa Dua'
    ws['A2'].font = Font(size=11)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Address line 2
    ws['A3'] = 'Tangerang, Banten 15812 - 0821 1213 4258'
    ws['A3'].font = Font(size=11)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Website & social media
    ws['A4'] = 'interiormapan.com - mapan.interiorr'
    ws['A4'].font = Font(size=11, color='e74c3c', bold=True)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add spacing
    ws.row_dimensions[5].height = 5
    
    # Date
    ws.merge_cells('A6:I6')
    ws['A6'] = f'Tanggal: {datetime.now().strftime("%d %B %Y, %H:%M")}'
    ws['A6'].font = Font(size=10, italic=True)
    ws['A6'].alignment = Alignment(horizontal='right')
    
    # Add spacing
    ws.row_dimensions[7].height = 10
    
    # ========== TABLE HEADERS ==========
    headers = ['No.', 'Nama Item', 'Sub Item', 'Deskripsi', 'Dimensi', 
               'Satuan Dimensi', 'Total Volume', 'Harga Dasar', 'Jumlah']
    
    header_row = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # ========== DATA ROWS ==========
    for idx, item in enumerate(st.session_state.items_data, 1):
        row_num = header_row + idx
        
        row_data = [
            idx,
            item['nama_item'],
            item['sub_item'],
            item['deskripsi'],
            item['dimensi'],
            item['satuan_dimensi'],
            item['total_volume'],
            item['harga_dasar'],
            item['jumlah']
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Alignment
            if col == 1:  # No
                cell.alignment = Alignment(horizontal='center')
            elif col in [7, 8, 9]:  # Numbers
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
            
            # Number formatting
            if col == 7:  # Total Volume
                cell.number_format = '0.00'
            elif col in [8, 9]:  # Prices
                cell.number_format = '#,##0'
    
    # ========== TOTAL ROW ==========
    total_row = header_row + len(st.session_state.items_data) + 1
    ws.merge_cells(f'A{total_row}:H{total_row}')
    ws[f'A{total_row}'] = 'TOTAL HARGA'
    ws[f'A{total_row}'].font = Font(bold=True, size=12)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{total_row}'].fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    
    total_harga = sum(item['jumlah'] for item in st.session_state.items_data)
    ws[f'I{total_row}'] = total_harga
    ws[f'I{total_row}'].font = Font(bold=True, size=12, color='c0392b')
    ws[f'I{total_row}'].number_format = '#,##0'
    ws[f'I{total_row}'].alignment = Alignment(horizontal='right')
    ws[f'I{total_row}'].fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    
    # ========== COLUMN WIDTHS ==========
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

# Main UI
# ========== HEADER SECTION ==========
header_col1, header_col2 = st.columns([1, 3])

with header_col1:
    # Logo placeholder
    st.markdown("""
        <div style='
            background-color: #c0392b;
            color: white;
            font-size: 48px;
            font-weight: bold;
            text-align: center;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 10px;
        '>
            EPD
        </div>
    """, unsafe_allow_html=True)

with header_col2:
    # Company info
    st.markdown("""
        <div style='
            background-color: white;
            border: 3px solid #c0392b;
            border-radius: 10px;
            padding: 15px;
            margin-bottom: 10px;
        '>
            <h2 style='color: #c0392b; margin: 0; font-size: 20px;'>PT. ENIGMA PRISMA DELAPAN</h2>
            <p style='margin: 5px 0; font-size: 13px; color: #2c3e50;'>
                Jl. Raya H. Abdullah No.56, Pakulonan Barat, Kelapa Dua<br>
                Tangerang, Banten 15812 - 0821 1213 4258<br>
                <span style='color: #e74c3c;'><b>interiormapan.com - mapan.interiorr</b></span>
            </p>
        </div>
    """, unsafe_allow_html=True)

# Separator
st.markdown("<hr style='border: 2px solid #c0392b; margin: 20px 0;'>", unsafe_allow_html=True)

# Title
st.title("🛋️ Furniture Price Generator")
st.markdown("**by Mapan** - Versi Web untuk semua device")

# Sidebar - Category selection
st.sidebar.header("Pilih Kategori")
category = st.sidebar.selectbox(
    "Kategori Furniture:",
    ["Kitchen Set", "Wardrobe", "Bed", "Backdrop", "Credenza", "Multi Cabinet", "Custom"]
)

# Main content area
col1, col2 = st.columns([2, 1])

with col1:
    st.header(f"📝 {category}")
    
    # KITCHEN SET
    if category == "Kitchen Set":
        with st.expander("🔽 Kabinet Bawah", expanded=True):
            kb_col1, kb_col2, kb_col3 = st.columns(3)
            with kb_col1:
                kb_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="kb_p")
            with kb_col2:
                kb_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="kb_t")
            with kb_col3:
                kb_tebal = st.number_input("Tebal (cm)", min_value=0.0, value=0.0, key="kb_tbl")
            
            kb_finishing = st.selectbox("Finishing Kabinet Bawah:", ["HPL", "Duco", "Kombinasi"], key="kb_f")
            if st.button("ℹ️ Info Finishing", key="kb_info"):
                st.info(MaterialDescriptions.get_description(kb_finishing))
            
            if st.button("➕ Tambah Kabinet Bawah", key="kb_add"):
                if kb_panjang > 0 and kb_tinggi > 0 and kb_tebal > 0:
                    total, area, unit_price = calculate_cabinet_price(kb_panjang, kb_tinggi, kb_tebal, kb_finishing)
                    st.session_state.items_data.append({
                        'nama_item': 'Kitchen Set',
                        'sub_item': 'Kabinet Bawah',
                        'deskripsi': MaterialDescriptions.get_description(kb_finishing),
                        'dimensi': f"{max(kb_panjang,100)} x {max(kb_tinggi,100)} x {kb_tebal}",
                        'satuan_dimensi': 'cm',
                        'total_volume': area,
                        'harga_dasar': unit_price,
                        'jumlah': total
                    })
                    st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                    st.rerun()
                else:
                    st.error("Masukkan semua dimensi!")
        
        with st.expander("🔼 Kabinet Atas", expanded=True):
            ka_col1, ka_col2, ka_col3 = st.columns(3)
            with ka_col1:
                ka_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="ka_p")
            with ka_col2:
                ka_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="ka_t")
            with ka_col3:
                ka_tebal = st.number_input("Tebal (cm)", min_value=0.0, value=0.0, key="ka_tbl")
            
            ka_finishing = st.selectbox("Finishing Kabinet Atas:", ["HPL", "Duco", "Kombinasi"], key="ka_f")
            if st.button("ℹ️ Info Finishing", key="ka_info"):
                st.info(MaterialDescriptions.get_description(ka_finishing))
            
            if st.button("➕ Tambah Kabinet Atas", key="ka_add"):
                if ka_panjang > 0 and ka_tinggi > 0 and ka_tebal > 0:
                    total, area, unit_price = calculate_cabinet_price(ka_panjang, ka_tinggi, ka_tebal, ka_finishing)
                    st.session_state.items_data.append({
                        'nama_item': 'Kitchen Set',
                        'sub_item': 'Kabinet Atas',
                        'deskripsi': MaterialDescriptions.get_description(ka_finishing),
                        'dimensi': f"{max(ka_panjang,100)} x {max(ka_tinggi,100)} x {ka_tebal}",
                        'satuan_dimensi': 'cm',
                        'total_volume': area,
                        'harga_dasar': unit_price,
                        'jumlah': total
                    })
                    st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                    st.rerun()
                else:
                    st.error("Masukkan semua dimensi!")
        
        with st.expander("📐 Top Table", expanded=True):
            tt_col1, tt_col2 = st.columns(2)
            with tt_col1:
                tt_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="tt_p")
            with tt_col2:
                tt_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="tt_t")
            
            tt_material = st.selectbox("Material:", ["Solid Surface", "Granit Alam", "Marmer", 
                                                      "Mirror Clear", "Bronze Mirror", "Keramik"], key="tt_m")
            if st.button("ℹ️ Info Material", key="tt_info"):
                st.info(MaterialDescriptions.get_description(tt_material))
            
            if st.button("➕ Tambah Top Table", key="tt_add"):
                if tt_panjang > 0 and tt_tinggi > 0:
                    total, area, unit_price = calculate_surface_price(tt_panjang, tt_tinggi, tt_material)
                    st.session_state.items_data.append({
                        'nama_item': 'Kitchen Set',
                        'sub_item': 'Top Table',
                        'deskripsi': MaterialDescriptions.get_description(tt_material),
                        'dimensi': f"{tt_panjang} x {tt_tinggi}",
                        'satuan_dimensi': 'cm',
                        'total_volume': area,
                        'harga_dasar': unit_price,
                        'jumlah': total
                    })
                    st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                    st.rerun()
                else:
                    st.error("Masukkan semua dimensi!")
        
        with st.expander("🔲 Backsplash", expanded=True):
            bs_col1, bs_col2 = st.columns(2)
            with bs_col1:
                bs_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="bs_p")
            with bs_col2:
                bs_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="bs_t")
            
            bs_material = st.selectbox("Material:", ["Solid Surface", "Granit Alam", "Marmer", 
                                                      "Mirror Clear", "Bronze Mirror", "Keramik"], key="bs_m")
            if st.button("ℹ️ Info Material", key="bs_info"):
                st.info(MaterialDescriptions.get_description(bs_material))
            
            if st.button("➕ Tambah Backsplash", key="bs_add"):
                if bs_panjang > 0 and bs_tinggi > 0:
                    total, area, unit_price = calculate_surface_price(bs_panjang, bs_tinggi, bs_material)
                    st.session_state.items_data.append({
                        'nama_item': 'Kitchen Set',
                        'sub_item': 'Backsplash',
                        'deskripsi': MaterialDescriptions.get_description(bs_material),
                        'dimensi': f"{bs_panjang} x {bs_tinggi}",
                        'satuan_dimensi': 'cm',
                        'total_volume': area,
                        'harga_dasar': unit_price,
                        'jumlah': total
                    })
                    st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                    st.rerun()
                else:
                    st.error("Masukkan semua dimensi!")
    
    # WARDROBE
    elif category == "Wardrobe":
        wr_col1, wr_col2, wr_col3 = st.columns(3)
        with wr_col1:
            wr_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="wr_p")
        with wr_col2:
            wr_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="wr_t")
        with wr_col3:
            wr_tebal = st.number_input("Tebal (cm)", min_value=0.0, value=0.0, key="wr_tbl")
        
        wr_finishing = st.selectbox("Finishing:", ["HPL", "Duco", "Kombinasi"], key="wr_f")
        if st.button("ℹ️ Info Finishing", key="wr_info"):
            st.info(MaterialDescriptions.get_description(wr_finishing))
        
        if st.button("➕ Tambah Wardrobe", key="wr_add"):
            if wr_panjang > 0 and wr_tinggi > 0 and wr_tebal > 0:
                total, area, unit_price = calculate_wardrobe_price(wr_panjang, wr_tinggi, wr_tebal, wr_finishing)
                st.session_state.items_data.append({
                    'nama_item': 'Wardrobe',
                    'sub_item': '-',
                    'deskripsi': MaterialDescriptions.get_description(wr_finishing),
                    'dimensi': f"{wr_panjang} x {wr_tinggi} x {wr_tebal}",
                    'satuan_dimensi': 'cm',
                    'total_volume': area,
                    'harga_dasar': unit_price,
                    'jumlah': total
                })
                st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                st.rerun()
            else:
                st.error("Masukkan semua dimensi!")
    
    # BED
    elif category == "Bed":
        bed_col1, bed_col2, bed_col3 = st.columns(3)
        with bed_col1:
            bed_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="bed_p")
        with bed_col2:
            bed_lebar = st.number_input("Lebar (cm)", min_value=0.0, value=0.0, key="bed_l")
        with bed_col3:
            bed_tinggi = st.number_input("Tinggi Sandaran (cm)", min_value=0.0, value=0.0, key="bed_ts")
        
        bed_finishing = st.selectbox("Finishing Sandaran:", ["HPL", "Duco", "Kombinasi"], key="bed_f")
        bed_material = st.selectbox("Material Dudukan:", ["Synthetic Leather", "Fabric"], key="bed_m")
        
        col_info1, col_info2 = st.columns(2)
        with col_info1:
            if st.button("ℹ️ Info Finishing", key="bed_info_f"):
                st.info(MaterialDescriptions.get_description(bed_finishing))
        with col_info2:
            if st.button("ℹ️ Info Material", key="bed_info_m"):
                st.info(MaterialDescriptions.get_description(bed_material))
        
        if st.button("➕ Tambah Bed", key="bed_add"):
            if bed_panjang > 0 and bed_lebar > 0 and bed_tinggi > 0:
                total, area_s, area_d, price_s, price_d = calculate_bed_price(
                    bed_panjang, bed_lebar, bed_tinggi, bed_finishing, bed_material
                )
                desc = f"Sandaran: {MaterialDescriptions.get_description(bed_finishing)}\nDudukan: {MaterialDescriptions.get_description(bed_material)}"
                st.session_state.items_data.append({
                    'nama_item': 'Bed',
                    'sub_item': '-',
                    'deskripsi': desc,
                    'dimensi': f"{bed_panjang} x {bed_lebar} x {bed_tinggi}",
                    'satuan_dimensi': 'cm',
                    'total_volume': area_s + area_d,
                    'harga_dasar': (price_s + price_d) / 2,
                    'jumlah': total
                })
                st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                st.rerun()
            else:
                st.error("Masukkan semua dimensi!")
    
    # BACKDROP
    elif category == "Backdrop":
        bd_col1, bd_col2 = st.columns(2)
        with bd_col1:
            bd_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="bd_p")
        with bd_col2:
            bd_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="bd_t")
        
        bd_type = st.selectbox("Type:", ["PVC", "Flat", "Tebal"], key="bd_type")
        
        # Keterangan type backdrop
        type_info = {
            'PVC': '🎨 **PVC Backdrop** → Material: PVC Panel Motif | Harga: Rp 1.500.000/m² | Finishing: Tidak perlu (sudah jadi)',
            'Flat': '📋 **Flat Panel (Tipis)** → Ketebalan: < 10cm | Multiplier: 1.0x | Finishing: Perlu (HPL/Duco/Kombinasi)',
            'Tebal': '📦 **Panel Tebal (3D)** → Ketebalan: ≥ 10cm | Multiplier: 1.3x (lebih mahal) | Finishing: Perlu (HPL/Duco/Kombinasi)'
        }
        st.info(type_info.get(bd_type, ''))
        
        bd_finishing = None
        if bd_type != 'PVC':
            bd_finishing = st.selectbox("Finishing:", ["HPL", "Duco", "Kombinasi"], key="bd_f")
            if st.button("ℹ️ Info Finishing", key="bd_info"):
                st.info(MaterialDescriptions.get_description(bd_finishing))
        
        if st.button("➕ Tambah Backdrop", key="bd_add"):
            if bd_panjang > 0 and bd_tinggi > 0:
                if bd_type == 'PVC':
                    area = (bd_panjang / 100) * (bd_tinggi / 100)
                    unit_price = 1500000
                    total = area * unit_price
                    desc = "PVC Panel Motif"
                elif bd_type == 'Flat':
                    total, area, unit_price = calculate_backdrop_price(bd_panjang, bd_tinggi, bd_finishing)
                    desc = MaterialDescriptions.get_description(bd_finishing)
                else:  # Tebal
                    base_total, area, base_price = calculate_backdrop_price(bd_panjang, bd_tinggi, bd_finishing)
                    unit_price = base_price * 1.3
                    total = base_total * 1.3
                    desc = MaterialDescriptions.get_description(bd_finishing) + " (Panel Tebal 3D)"
                
                st.session_state.items_data.append({
                    'nama_item': 'Backdrop',
                    'sub_item': bd_type,
                    'deskripsi': desc,
                    'dimensi': f"{bd_panjang} x {bd_tinggi}",
                    'satuan_dimensi': 'cm',
                    'total_volume': area,
                    'harga_dasar': unit_price,
                    'jumlah': total
                })
                st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                st.rerun()
            else:
                st.error("Masukkan semua dimensi!")
    
    # CREDENZA
    elif category == "Credenza":
        cr_col1, cr_col2, cr_col3 = st.columns(3)
        with cr_col1:
            cr_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="cr_p")
        with cr_col2:
            cr_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="cr_t")
        with cr_col3:
            cr_tebal = st.number_input("Tebal (cm)", min_value=0.0, value=0.0, key="cr_tbl")
        
        cr_finishing = st.selectbox("Finishing:", ["HPL", "Duco", "Kombinasi"], key="cr_f")
        if st.button("ℹ️ Info Finishing", key="cr_info"):
            st.info(MaterialDescriptions.get_description(cr_finishing))
        
        if st.button("➕ Tambah Credenza", key="cr_add"):
            if cr_panjang > 0 and cr_tinggi > 0 and cr_tebal > 0:
                total, area, unit_price = calculate_credenza_price(cr_panjang, cr_tinggi, cr_tebal, cr_finishing)
                st.session_state.items_data.append({
                    'nama_item': 'Credenza',
                    'sub_item': '-',
                    'deskripsi': MaterialDescriptions.get_description(cr_finishing),
                    'dimensi': f"{cr_panjang} x {cr_tinggi} x {cr_tebal}",
                    'satuan_dimensi': 'cm',
                    'total_volume': area,
                    'harga_dasar': unit_price,
                    'jumlah': total
                })
                st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                st.rerun()
            else:
                st.error("Masukkan semua dimensi!")
    
    # MULTI CABINET
    elif category == "Multi Cabinet":
        mc_col1, mc_col2, mc_col3 = st.columns(3)
        with mc_col1:
            mc_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="mc_p")
        with mc_col2:
            mc_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="mc_t")
        with mc_col3:
            mc_tebal = st.number_input("Tebal (cm)", min_value=0.0, value=0.0, key="mc_tbl")
        
        mc_finishing = st.selectbox("Finishing:", ["HPL", "Duco", "Kombinasi"], key="mc_f")
        if st.button("ℹ️ Info Finishing", key="mc_info"):
            st.info(MaterialDescriptions.get_description(mc_finishing))
        
        if st.button("➕ Tambah Multi Cabinet", key="mc_add"):
            if mc_panjang > 0 and mc_tinggi > 0 and mc_tebal > 0:
                total, area, unit_price = calculate_multi_cabinet_price(mc_panjang, mc_tinggi, mc_tebal, mc_finishing)
                st.session_state.items_data.append({
                    'nama_item': 'Multi Cabinet',
                    'sub_item': '-',
                    'deskripsi': MaterialDescriptions.get_description(mc_finishing),
                    'dimensi': f"{mc_panjang} x {mc_tinggi} x {mc_tebal}",
                    'satuan_dimensi': 'cm',
                    'total_volume': area,
                    'harga_dasar': unit_price,
                    'jumlah': total
                })
                st.success(f"✅ Ditambahkan: Rp {total:,.0f}")
                st.rerun()
            else:
                st.error("Masukkan semua dimensi!")
    
    # CUSTOM
    elif category == "Custom":
        cust_nama = st.text_input("Nama Item Custom:", key="cust_nama")
        
        cust_col1, cust_col2, cust_col3 = st.columns(3)
        with cust_col1:
            cust_panjang = st.number_input("Panjang (cm)", min_value=0.0, value=0.0, key="cust_p")
        with cust_col2:
            cust_lebar = st.number_input("Lebar (cm)", min_value=0.0, value=0.0, key="cust_l")
        with cust_col3:
            cust_tinggi = st.number_input("Tinggi (cm)", min_value=0.0, value=0.0, key="cust_t")
        
        cust_method = st.selectbox("Metode Hitung:", 
            ["Meter Lari", "Meter Persegi", "Side Area X", "Side Area Y", "Side Area Z"],
            key="cust_method")
        
        # Keterangan metode
        method_info = {
            'Meter Lari': '📏 **Meter Lari** → Rumus: Panjang ÷ 100 | Contoh: 300cm = 3.0 meter | Untuk: kitchen set, backsplash, list',
            'Meter Persegi': '📐 **Meter Persegi** → Rumus: (P × L) ÷ 10,000 | Contoh: 200×100cm = 2.0 m² | Untuk: panel datar',
            'Side Area X': '🔴 **Side Area X (Sisi Samping)** → Rumus: (L × T) ÷ 10,000 | Contoh: 60×200cm = 1.2 m² | Untuk: sisi kiri/kanan',
            'Side Area Y': '🟢 **Side Area Y (Depan/Belakang)** → Rumus: (P × T) ÷ 10,000 | Contoh: 300×85cm = 2.55 m² | Untuk: kabinet, backdrop',
            'Side Area Z': '🔵 **Side Area Z (Atas/Bawah)** → Rumus: (P × L) ÷ 10,000 | Contoh: 150×80cm = 1.2 m² | Untuk: permukaan atas'
        }
        st.info(method_info.get(cust_method, ''))
        
        cust_unit_price = st.number_input("Harga Satuan (Rp/unit):", min_value=0.0, value=0.0, key="cust_up")
        
        if st.button("➕ Tambah Custom Item", key="cust_add"):
            if cust_nama and cust_panjang > 0 and cust_unit_price > 0:
                # Hitung volume berdasarkan metode
                if cust_method == 'Meter Lari':
                    volume = cust_panjang / 100
                    dimensi = f"{cust_panjang}"
                elif cust_method == 'Meter Persegi':
                    panjang = max(cust_panjang, 100)
                    lebar = max(cust_lebar, 100)
                    volume = (panjang / 100) * (lebar / 100)
                    dimensi = f"{panjang} x {lebar}"
                elif cust_method == 'Side Area X':
                    volume = (cust_lebar / 100) * (cust_tinggi / 100)
                    dimensi = f"{cust_panjang} x {cust_lebar} x {cust_tinggi}"
                elif cust_method == 'Side Area Y':
                    volume = (cust_panjang / 100) * (cust_tinggi / 100)
                    dimensi = f"{cust_panjang} x {cust_lebar} x {cust_tinggi}"
                else:  # Side Area Z
                    volume = (cust_panjang / 100) * (cust_lebar / 100)
                    dimensi = f"{cust_panjang} x {cust_lebar} x {cust_tinggi}"
                
                total = volume * cust_unit_price
                
                st.session_state.items_data.append({
                    'nama_item': 'Custom',
                    'sub_item': cust_nama,
                    'deskripsi': f"Metode: {cust_method}",
                    'dimensi': dimensi,
                    'satuan_dimensi': 'cm',
                    'total_volume': volume,
                    'harga_dasar': cust_unit_price,
                    'jumlah': total
                })
                st.success(f"✅ Ditambahkan: {cust_nama} - Rp {total:,.0f}")
                st.rerun()
            else:
                st.error("Lengkapi nama, panjang, dan harga satuan!")

# Right column - Items list and actions
with col2:
    st.header("📋 Daftar Item")
    
    if st.session_state.items_data:
        # Display items
        total_harga = 0
        for idx, item in enumerate(st.session_state.items_data):
            with st.container():
                st.markdown(f"**{idx+1}. {item['nama_item']}** - {item['sub_item']}")
                st.caption(f"💰 Rp {item['jumlah']:,.0f}")
                st.caption(f"📐 {item['dimensi']} cm")
                if st.button("🗑️ Hapus", key=f"del_{idx}"):
                    st.session_state.items_data.pop(idx)
                    st.rerun()
                st.divider()
            total_harga += item['jumlah']
        
        # Total
        st.markdown("---")
        st.markdown(f"### 💵 Total: Rp {total_harga:,.0f}")
        
        # Discount
        st.markdown("---")
        discount = st.slider("Diskon (%)", 0, 50, 0, 5)
        if discount > 0:
            discounted_total = total_harga * (1 - discount/100)
            st.markdown(f"**Setelah diskon {discount}%:**")
            st.markdown(f"### 💰 Rp {discounted_total:,.0f}")
        
        # Export
        st.markdown("---")
        if st.button("📥 Export ke Excel", use_container_width=True):
            excel_file = export_to_excel()
            if excel_file:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel_file,
                    file_name=f"furniture_price_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        
        # Clear all
        if st.button("🗑️ Hapus Semua Item", use_container_width=True):
            st.session_state.items_data = []
            st.rerun()
    else:
        st.info("Belum ada item. Tambahkan item dari form di sebelah kiri.")

# Footer
st.sidebar.markdown("---")
st.sidebar.markdown("### 📱 Akses dari Device Manapun")
st.sidebar.info("""
Aplikasi ini bisa diakses dari:
- 💻 Windows
- 🤖 Android
- 🍎 iOS
- 🖥️ Mac
- 📱 Tablet

Cukup buka di browser!
""")
