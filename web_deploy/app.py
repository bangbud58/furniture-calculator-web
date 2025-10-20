from flask import Flask, render_template, request, jsonify, session, send_file
import json
import os
from datetime import datetime
import tempfile
import requests
import urllib.parse
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.utils import secure_filename
import uuid

# Copy data dari module yang sudah ada
try:
    import sys
    sys.path.append('..')
    from info_combo import InfoComboBox
    from material_descriptions import MaterialDescriptions
except ImportError:
    # Fallback jika tidak bisa import dari parent directory
    class InfoComboBox:
        @staticmethod
        def get_info(category, sub_category):
            return f"Info untuk {category} - {sub_category}"
    
    class MaterialDescriptions:
        FINISHING_DETAILS = {
            'Tacosheet': 'Luar tacosheet, dalam melamninto, ekonomis (Rp 2,3jt/m²)',
            'HPL Low': 'Luar HPL, dalam melamninto, motif terbatas (Rp 2,3jt/m²)',
            'HPL Mid': 'Luar HPL, dalam melamninto, motif standar (Rp 2,8jt/m²)',
            'HPL High': 'Luar+dalam full HPL, semua motif tersedia (Rp 3,2jt/m²)',
            'Duco': 'Cat duco premium, finishing halus (Rp 5,5jt/m²)',
            'Kombinasi': 'Kombinasi HPL + duco (Rp 4,7jt/m²)'
        }

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'furniture_price_generator_2025_prod')

# Configuration for serverless
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Use /tmp for temporary files in serverless environment
UPLOAD_FOLDER = '/tmp' if os.environ.get('VERCEL') else 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Ensure upload folder exists (only if not in serverless)
if not os.environ.get('VERCEL'):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

class FurniturePriceCalculator:
    """Core calculation logic"""
    
    @staticmethod
    def get_price_per_m3(category, finishing):
        """Get price per m³ based on category and finishing"""
        # Base prices - EXACT SAME AS DESKTOP VERSION
        prices = {
            'Tacosheet': 2300000,    # Same as desktop
            'HPL Low': 2300000,      # Same as desktop HPL Low
            'HPL Mid': 2800000,      # Same as desktop HPL Mid  
            'HPL High': 3200000,     # Same as desktop HPL High
            'Duco': 5500000,         # Same as desktop
            'Kombinasi': 4700000     # Same as desktop
        }
        
        # Category-specific adjustments for bed and backdrop
        if category in ['Bed Frame']:
            # Bed uses old pricing system
            if finishing == 'HPL':
                return 3000000  # Old HPL price for bed
            elif finishing == 'Duco':
                return 5500000
            # No Kombinasi for bed
        
        if category in ['Backdrop Panel']:
            # Backdrop uses old system with type variations
            base_price = prices.get(finishing, 2300000)
            return base_price
        
        return prices.get(finishing, 2300000)
    
    @staticmethod
    def calculate_total(items_data, discount_percent=0, tax_percent=11, use_tax=True):
        """Calculate total with discount and tax"""
        if not items_data:
            return {
                'subtotal': 0,
                'discount_amount': 0,
                'after_discount': 0,
                'tax_amount': 0,
                'grand_total': 0,
                'item_count': 0
            }
        
        subtotal = sum(float(item['jumlah']) for item in items_data)
        discount_amount = subtotal * (discount_percent / 100)
        after_discount = subtotal - discount_amount
        
        if use_tax:
            tax_amount = after_discount * (tax_percent / 100)
            grand_total = after_discount + tax_amount
        else:
            tax_amount = 0
            grand_total = after_discount
        
        return {
            'subtotal': subtotal,
            'discount_amount': discount_amount,
            'after_discount': after_discount,
            'tax_amount': tax_amount,
            'grand_total': grand_total,
            'item_count': len(items_data)
        }

def init_session():
    """Initialize session data"""
    if 'items' not in session:
        session['items'] = []
    if 'customer_info' not in session:
        session['customer_info'] = {
            'name': '',
            'address': '',
            'phone': ''
        }
    if 'pricing_info' not in session:
        session['pricing_info'] = {
            'discount_percent': 0.0,
            'tax_percent': 11.0,
            'use_tax': True
        }

@app.route('/')
def index():
    """Main page"""
    init_session()
    
    # Category sub items mapping
    category_subs = {
        'Kitchen Set KB': ['Kitchen Bawah'],
        'Kitchen Set KA': ['Kitchen Atas'],
        'Wardrobe': ['2 Pintu', '3 Pintu', '4 Pintu'],
        'Bed Frame': ['Single', 'Double', 'Queen', 'King'],
        'Backdrop Panel': ['TV Wall', 'Headboard', 'Partition'],
        'Credenza': ['Credenza 120', 'Credenza 150', 'Credenza 180'],
        'Multi Cabinet': ['2 Level', '3 Level', '4 Level'],
        'Custom': ['Custom Item']
    }
    
    # Finishing options by category
    finishing_options = {
        'Kitchen Set KB': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Kitchen Set KA': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Wardrobe': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Bed Frame': ['HPL', 'Duco'],  # Old system, no new tiers
        'Backdrop Panel': ['HPL', 'Duco', 'Kombinasi'],  # Old system
        'Credenza': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Multi Cabinet': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Custom': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi']
    }
    
    # Calculate current total
    total_info = FurniturePriceCalculator.calculate_total(
        session['items'],
        session['pricing_info']['discount_percent'],
        session['pricing_info']['tax_percent'],
        session['pricing_info']['use_tax']
    )
    
    return render_template('index.html', 
                         category_subs=category_subs,
                         finishing_options=finishing_options,
                         items=session['items'],
                         customer_info=session['customer_info'],
                         pricing_info=session['pricing_info'],
                         total_info=total_info)

@app.route('/add_item', methods=['POST'])
def add_item():
    """Add item to the list"""
    try:
        init_session()
        
        data = request.get_json()
        
        # Validate input
        panjang = float(data['panjang'])
        lebar = float(data['lebar'])
        tinggi = float(data['tinggi'])
        
        if panjang <= 0 or lebar <= 0 or tinggi <= 0:
            return jsonify({'success': False, 'message': 'Dimensi harus lebih besar dari 0!'})
        
        # Calculate volume and price
        volume = panjang * lebar * tinggi
        unit_price = FurniturePriceCalculator.get_price_per_m3(
            data['category'], data['finishing']
        )
        total_price = volume * unit_price
        
        # Create item data
        item_data = {
            'id': str(uuid.uuid4()),  # Unique ID for web
            'nama_item': data['category'],
            'sub_item': data['sub_item'],
            'deskripsi': f"Finishing {data['finishing']}",
            'dimensi': f"{panjang}×{lebar}×{tinggi}",
            'satuan_dimensi': 'm³',
            'total_volume': volume,
            'harga_dasar': unit_price,
            'jumlah': total_price
        }
        
        # Add to session
        session['items'].append(item_data)
        session.modified = True
        
        # Calculate new total
        total_info = FurniturePriceCalculator.calculate_total(
            session['items'],
            session['pricing_info']['discount_percent'],
            session['pricing_info']['tax_percent'],
            session['pricing_info']['use_tax']
        )
        
        return jsonify({
            'success': True, 
            'message': 'Item berhasil ditambahkan!',
            'item': item_data,
            'total_info': total_info
        })
        
    except (ValueError, KeyError) as e:
        return jsonify({'success': False, 'message': f'Input tidak valid: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/delete_item', methods=['POST'])
def delete_item():
    """Delete item from the list"""
    try:
        init_session()
        
        data = request.get_json()
        item_id = data['item_id']
        
        # Find and remove item
        session['items'] = [item for item in session['items'] if item.get('id') != item_id]
        session.modified = True
        
        # Calculate new total
        total_info = FurniturePriceCalculator.calculate_total(
            session['items'],
            session['pricing_info']['discount_percent'],
            session['pricing_info']['tax_percent'],
            session['pricing_info']['use_tax']
        )
        
        return jsonify({
            'success': True, 
            'message': 'Item berhasil dihapus!',
            'total_info': total_info
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/clear_all', methods=['POST'])
def clear_all():
    """Clear all items"""
    init_session()
    session['items'] = []
    session.modified = True
    
    total_info = FurniturePriceCalculator.calculate_total([], 0, 11, True)
    
    return jsonify({
        'success': True, 
        'message': 'Semua item berhasil dihapus!',
        'total_info': total_info
    })

@app.route('/update_customer', methods=['POST'])
def update_customer():
    """Update customer information"""
    try:
        init_session()
        
        data = request.get_json()
        
        session['customer_info'] = {
            'name': data.get('name', ''),
            'address': data.get('address', ''),
            'phone': data.get('phone', '')
        }
        session.modified = True
        
        return jsonify({'success': True, 'message': 'Info customer berhasil disimpan!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/update_pricing', methods=['POST'])
def update_pricing():
    """Update pricing information"""
    try:
        init_session()
        
        data = request.get_json()
        
        session['pricing_info'] = {
            'discount_percent': float(data.get('discount_percent', 0)),
            'tax_percent': float(data.get('tax_percent', 11)),
            'use_tax': bool(data.get('use_tax', True))
        }
        session.modified = True
        
        # Calculate new total
        total_info = FurniturePriceCalculator.calculate_total(
            session['items'],
            session['pricing_info']['discount_percent'],
            session['pricing_info']['tax_percent'],
            session['pricing_info']['use_tax']
        )
        
        tax_status = f"Pajak: {session['pricing_info']['tax_percent']}%" if session['pricing_info']['use_tax'] else "Pajak: Tidak Digunakan"
        
        return jsonify({
            'success': True, 
            'message': f"Diskon: {session['pricing_info']['discount_percent']}% | {tax_status}",
            'total_info': total_info
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/get_total')
def get_total():
    """Get total calculation"""
    init_session()
    
    total_info = FurniturePriceCalculator.calculate_total(
        session['items'],
        session['pricing_info']['discount_percent'],
        session['pricing_info']['tax_percent'],
        session['pricing_info']['use_tax']
    )
    
    return jsonify(total_info)

@app.route('/save_project', methods=['POST'])
def save_project():
    """Save project to JSON and return as download"""
    try:
        init_session()
        
        if not session['items']:
            return jsonify({'success': False, 'message': 'Tidak ada data untuk disimpan!'})
        
        # Generate filename
        now = datetime.now()
        year = now.strftime('%Y')
        roman_months = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII']
        month_roman = roman_months[now.month - 1]
        sequence = now.strftime('%d%H%M')
        customer_name = session['customer_info']['name'].strip() if session['customer_info']['name'] else "NoName"
        customer_name = "".join(c for c in customer_name if c.isalnum() or c in (' ', '-', '_')).strip()
        
        filename = f'MPN{year}{month_roman}{sequence}-{customer_name}.json'
        
        # Create project data
        project_data = {
            'customer_info': session['customer_info'],
            'pricing_info': session['pricing_info'],
            'items': session['items'],
            'saved_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8')
        json.dump(project_data, temp_file, indent=2, ensure_ascii=False)
        temp_file.close()
        
        return jsonify({
            'success': True, 
            'message': f'Project siap didownload: {filename}',
            'download_url': f'/download_project/{os.path.basename(temp_file.name)}',
            'filename': filename
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/download_project/<temp_filename>')
def download_project(temp_filename):
    """Download saved project file"""
    try:
        temp_file_path = os.path.join(tempfile.gettempdir(), temp_filename)
        
        if not os.path.exists(temp_file_path):
            return jsonify({'error': 'File tidak ditemukan'}), 404
        
        # Determine actual filename from session
        customer_name = session.get('customer_info', {}).get('name', 'NoName').strip()
        if not customer_name:
            customer_name = 'NoName'
        customer_name = "".join(c for c in customer_name if c.isalnum() or c in (' ', '-', '_')).strip()
        
        now = datetime.now()
        year = now.strftime('%Y')
        roman_months = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII']
        month_roman = roman_months[now.month - 1]
        sequence = now.strftime('%d%H%M')
        
        download_filename = f'MPN{year}{month_roman}{sequence}-{customer_name}.json'
        
        return send_file(temp_file_path, 
                        as_attachment=True, 
                        download_name=download_filename,
                        mimetype='application/json')
        
    except Exception as e:
        return jsonify({'error': f'Error downloading file: {str(e)}'}), 500

@app.route('/load_project', methods=['POST'])
def load_project():
    """Load project from uploaded JSON"""
    try:
        init_session()
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file selected!'})
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected!'})
        
        if file and file.filename.lower().endswith('.json'):
            try:
                # Read and parse JSON
                content = file.read().decode('utf-8')
                project_data = json.loads(content)
                
                # Load data to session
                session['customer_info'] = project_data.get('customer_info', {
                    'name': '', 'address': '', 'phone': ''
                })
                session['pricing_info'] = project_data.get('pricing_info', {
                    'discount_percent': 0.0, 'tax_percent': 11.0, 'use_tax': True
                })
                
                # Add unique IDs to items if not present
                items = project_data.get('items', [])
                for item in items:
                    if 'id' not in item:
                        item['id'] = str(uuid.uuid4())
                
                session['items'] = items
                session.modified = True
                
                # Calculate total
                total_info = FurniturePriceCalculator.calculate_total(
                    session['items'],
                    session['pricing_info']['discount_percent'],
                    session['pricing_info']['tax_percent'],
                    session['pricing_info']['use_tax']
                )
                
                return jsonify({
                    'success': True, 
                    'message': f'Project berhasil dimuat: {file.filename}',
                    'customer_info': session['customer_info'],
                    'pricing_info': session['pricing_info'],
                    'items': session['items'],
                    'total_info': total_info
                })
                
            except json.JSONDecodeError:
                return jsonify({'success': False, 'message': 'File JSON tidak valid!'})
            
        else:
            return jsonify({'success': False, 'message': 'Please upload a JSON file!'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/export_excel', methods=['POST'])
def export_excel():
    """Export items to Excel with EXACT SAME formatting as desktop version"""
    items = session.get('items', [])
    customer_info = session.get('customer_info', {})
    pricing_info = session.get('pricing_info', {})
    
    if not items:
        return jsonify({'success': False, 'message': 'Tidak ada data untuk diekspor'})
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.drawing.image import Image as XLImage
        from datetime import datetime
        import tempfile
        import os
    except ImportError:
        return jsonify({'success': False, 'message': 'Library openpyxl belum terpasang'})

    wb = Workbook()
    ws = wb.active
    ws.title = 'Daftar Harga'

    # ========== LOGO & COMPANY HEADER - SAME AS DESKTOP ==========
    # Add logo if exists
    logo_path = os.path.join(app.static_folder, 'images', 'logo.png')
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            # Set ukuran logo H=2.7cm (~102 pixels), W=7cm (~264 pixels) - SAME AS DESKTOP
            img.width = 264  # 7cm
            img.height = 102  # 2.7cm
            # Posisi logo di A1 - SAME AS DESKTOP
            ws.add_image(img, 'A1')
            # Set tinggi row untuk logo - SAME AS DESKTOP
            ws.row_dimensions[1].height = 60
            ws.row_dimensions[2].height = 20
        except Exception as e:
            print(f"Gagal menambahkan logo: {e}")
    
    # ========== COMPANY INFO (rata kanan, sejajar dengan logo) - SAME AS DESKTOP ==========
    ws.merge_cells('D1:I1')
    ws.merge_cells('D2:I2')
    ws.merge_cells('D3:I3')
    ws.merge_cells('D4:I4')
    
    # Company name - SAME AS DESKTOP
    ws['D1'] = 'PT. ENIGMA PRISMA DELAPAN'
    ws['D1'].font = Font(bold=True, size=14, color='c0392b')
    ws['D1'].alignment = Alignment(horizontal='right', vertical='center')
    
    # Address lines - SAME AS DESKTOP
    ws['D2'] = 'Jl. Raya H. Abdullah No.56, Pakulonan Barat, Kelapa Dua'
    ws['D2'].font = Font(size=9)
    ws['D2'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws['D3'] = 'Tangerang, Banten 15812 - 0821 1213 4258'
    ws['D3'].font = Font(size=9)
    ws['D3'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws['D4'] = 'interiormapan.com - mapan.interiorr'
    ws['D4'].font = Font(size=9, color='e74c3c', bold=True)
    ws['D4'].alignment = Alignment(horizontal='right', vertical='center')
    
    # Add spacing - SAME AS DESKTOP
    ws.row_dimensions[5].height = 5
    
    # Date - SAME AS DESKTOP
    ws.merge_cells('A6:I6')
    ws['A6'] = f'Tanggal: {datetime.now().strftime("%d %B %Y, %H:%M")}'
    ws['A6'].font = Font(size=10, italic=True)
    ws['A6'].alignment = Alignment(horizontal='right')
    
    ws.row_dimensions[7].height = 10
    
    # ========== TABLE HEADERS - SAME AS DESKTOP ==========
    headers = ['No.', 'Nama Item', 'Sub Item', 'Deskripsi', 'Dimensi', 
              'Satuan Dimensi', 'Total Volume', 'Harga Dasar', 'Jumlah']
    
    header_row = 8
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # ========== DATA ROWS - SAME FORMAT AS DESKTOP ==========
    for idx, item in enumerate(items, start=1):
        row = [
            idx,
            item['category'],
            item['sub_item'],
            item.get('deskripsi', f"{item['finishing']} finishing"),
            f"{item['panjang']}×{item['lebar']}×{item['tinggi']}",
            'meter',
            float(item['volume']),
            float(item['harga']) / float(item['volume']) if float(item['volume']) > 0 else 0,
            float(item['harga'])
        ]
        
        row_num = header_row + idx
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = value
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Harga (H dan I) rata kanan, sisanya center - SAME AS DESKTOP
            if col_idx in [8, 9]:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Number formatting - SAME AS DESKTOP
    for r in range(header_row + 1, header_row + 1 + len(items)):
        ws[f'G{r}'].number_format = '0.00'
        ws[f'H{r}'].number_format = '"Rp "#,##0'
        ws[f'I{r}'].number_format = '"Rp "#,##0'
    
    # ========== TOTAL CALCULATIONS - SAME AS DESKTOP ==========
    total_row = header_row + len(items) + 2
    subtotal = sum(float(item['harga']) for item in items)
    
    # Subtotal - SAME AS DESKTOP
    ws.merge_cells(f'A{total_row}:H{total_row}')
    ws[f'A{total_row}'] = 'SUBTOTAL'
    ws[f'A{total_row}'].font = Font(bold=True, size=11)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'I{total_row}'] = subtotal
    ws[f'I{total_row}'].font = Font(bold=True, size=11)
    ws[f'I{total_row}'].number_format = '"Rp "#,##0'
    ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    
    total_row += 1
    
    # Discount calculations - SAME AS DESKTOP
    discount_percent = pricing_info.get('discount_percent', 0)
    if discount_percent > 0:
        discount_amount = subtotal * (discount_percent / 100)
        ws.merge_cells(f'A{total_row}:H{total_row}')
        ws[f'A{total_row}'] = f'DISKON ({discount_percent}%)'
        ws[f'A{total_row}'].font = Font(size=10, color='e74c3c')
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'I{total_row}'] = -discount_amount
        ws[f'I{total_row}'].font = Font(size=10, color='e74c3c')
        ws[f'I{total_row}'].number_format = '"Rp "#,##0'
        ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        total_row += 1
        after_discount = subtotal - discount_amount
    else:
        after_discount = subtotal
    
    # Tax calculations - SAME AS DESKTOP
    enable_tax = pricing_info.get('enable_tax', False)
    if enable_tax:
        tax_percent = pricing_info.get('tax_percent', 11)
        tax_amount = after_discount * (tax_percent / 100)
        ws.merge_cells(f'A{total_row}:H{total_row}')
        ws[f'A{total_row}'] = f'PAJAK/PPN ({tax_percent}%)'
        ws[f'A{total_row}'].font = Font(size=10)
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'I{total_row}'] = tax_amount
        ws[f'I{total_row}'].font = Font(size=10)
        ws[f'I{total_row}'].number_format = '"Rp "#,##0'
        ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        total_row += 1
        grand_total = after_discount + tax_amount
    else:
        grand_total = after_discount
    
    # Grand Total (warna merah) - SAME AS DESKTOP
    ws.merge_cells(f'A{total_row}:H{total_row}')
    ws[f'A{total_row}'] = 'GRAND TOTAL'
    ws[f'A{total_row}'].font = Font(bold=True, size=13, color='FFFFFF')
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{total_row}'].fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
    ws[f'I{total_row}'] = grand_total
    ws[f'I{total_row}'].font = Font(bold=True, size=13, color='FFFFFF')
    ws[f'I{total_row}'].number_format = '"Rp "#,##0'
    ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'I{total_row}'].fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
    
    # ========== CUSTOMER INFO - SAME AS DESKTOP ==========
    customer_name = customer_info.get('nama', '')
    customer_address = customer_info.get('alamat', '')
    customer_phone = customer_info.get('telepon', '')
    
    if customer_name or customer_address or customer_phone:
        total_row += 2
        ws.merge_cells(f'A{total_row}:I{total_row}')
        ws[f'A{total_row}'] = '👤 INFORMASI CUSTOMER'
        ws[f'A{total_row}'].font = Font(bold=True, size=11, color='2c3e50')
        ws[f'A{total_row}'].fill = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')
        ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        total_row += 1
        if customer_name:
            ws.merge_cells(f'A{total_row}:I{total_row}')
            ws[f'A{total_row}'] = f'Nama: {customer_name}'
            ws[f'A{total_row}'].font = Font(size=10)
            ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
            total_row += 1
        
        if customer_address:
            ws.merge_cells(f'A{total_row}:I{total_row}')
            ws[f'A{total_row}'] = f'Alamat: {customer_address}'
            ws[f'A{total_row}'].font = Font(size=10)
            ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
            total_row += 1
        
        if customer_phone:
            ws.merge_cells(f'A{total_row}:I{total_row}')
            ws[f'A{total_row}'] = f'Telepon: {customer_phone}'
            ws[f'A{total_row}'].font = Font(size=10)
            ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')

    # ========== COLUMN WIDTHS - SAME AS DESKTOP ==========
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18

    # Generate filename - SAME FORMAT AS DESKTOP
    now = datetime.now()
    year = now.strftime('%Y')
    roman_months = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII']
    month_roman = roman_months[now.month - 1]
    sequence = now.strftime('%d%H%M')
    
    customer_name_clean = customer_name.strip() if customer_name else "NoName"
    customer_name_clean = "".join(c for c in customer_name_clean if c.isalnum() or c in (' ', '-', '_')).strip()
    
    filename = f'MPN{year}{month_roman}{sequence}-{customer_name_clean}.xlsx'
    
    # Save to temporary file
    temp_file = os.path.join(tempfile.gettempdir(), filename)
    wb.save(temp_file)
    
    return send_file(temp_file, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/get_finishing_options/<category>')
def get_finishing_options(category):
    """Get finishing options for specific category"""
    finishing_options = {
        'Kitchen Set KB': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Kitchen Set KA': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Wardrobe': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Bed Frame': ['HPL', 'Duco'],  # Old system
        'Backdrop Panel': ['HPL', 'Duco', 'Kombinasi'],  # Old system
        'Credenza': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Multi Cabinet': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'],
        'Custom': ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi']
    }
    
    return jsonify(finishing_options.get(category, ['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi']))

def send_whatsapp_message(phone_number, message):
    """
    Kirim pesan WhatsApp menggunakan WhatsApp Web API atau service
    """
    try:
        # Format nomor telepon (hapus karakter non-digit)
        clean_phone = ''.join(filter(str.isdigit, phone_number))
        
        # Pastikan dimulai dengan 62 (Indonesia)
        if clean_phone.startswith('0'):
            clean_phone = '62' + clean_phone[1:]
        elif not clean_phone.startswith('62'):
            clean_phone = '62' + clean_phone
        
        # Encode pesan untuk URL
        encoded_message = urllib.parse.quote(message)
        
        # Buat URL WhatsApp Web
        whatsapp_url = f"https://api.whatsapp.com/send?phone={clean_phone}&text={encoded_message}"
        
        return {
            'success': True,
            'whatsapp_url': whatsapp_url,
            'phone_number': clean_phone,
            'message': 'WhatsApp URL generated successfully'
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@app.route('/export_excel_whatsapp', methods=['POST'])
def export_excel_whatsapp():
    """Export Excel (DESKTOP-LIKE) dan kirim via WhatsApp"""
    try:
        data = request.get_json()
        if not data or 'items' not in data:
            return jsonify({'success': False, 'error': 'Data tidak valid'}), 400

        items_data = data['items']
        customer_data = data.get('customer', {})
        pricing_info = data.get('pricing', {}) or {}
        totals_data = data.get('totals', {}) or {}

        if not items_data:
            return jsonify({'success': False, 'error': 'Tidak ada item untuk diekspor'}), 400

        # Build workbook with the EXACT SAME layout as export_excel
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.drawing.image import Image as XLImage
        from datetime import datetime
        import tempfile, os

        wb = Workbook()
        ws = wb.active
        ws.title = 'Daftar Harga'

        # Logo
        logo_path = os.path.join(app.static_folder, 'images', 'logo.png')
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 264
                img.height = 102
                ws.add_image(img, 'A1')
                ws.row_dimensions[1].height = 60
                ws.row_dimensions[2].height = 20
            except Exception as e:
                print(f"Gagal menambahkan logo: {e}")

        # Company info (right-aligned)
        ws.merge_cells('D1:I1')
        ws.merge_cells('D2:I2')
        ws.merge_cells('D3:I3')
        ws.merge_cells('D4:I4')
        ws['D1'] = 'PT. ENIGMA PRISMA DELAPAN'
        ws['D1'].font = Font(bold=True, size=14, color='c0392b')
        ws['D1'].alignment = Alignment(horizontal='right', vertical='center')
        ws['D2'] = 'Jl. Raya H. Abdullah No.56, Pakulonan Barat, Kelapa Dua'
        ws['D2'].font = Font(size=9)
        ws['D2'].alignment = Alignment(horizontal='right', vertical='center')
        ws['D3'] = 'Tangerang, Banten 15812 - 0821 1213 4258'
        ws['D3'].font = Font(size=9)
        ws['D3'].alignment = Alignment(horizontal='right', vertical='center')
        ws['D4'] = 'interiormapan.com - mapan.interiorr'
        ws['D4'].font = Font(size=9, color='e74c3c', bold=True)
        ws['D4'].alignment = Alignment(horizontal='right', vertical='center')

        ws.row_dimensions[5].height = 5
        ws.merge_cells('A6:I6')
        ws['A6'] = f'Tanggal: {datetime.now().strftime("%d %B %Y, %H:%M")}'
        ws['A6'].font = Font(size=10, italic=True)
        ws['A6'].alignment = Alignment(horizontal='right')
        ws.row_dimensions[7].height = 10

        # Table headers
        headers = ['No.', 'Nama Item', 'Sub Item', 'Deskripsi', 'Dimensi', 'Satuan Dimensi', 'Total Volume', 'Harga Dasar', 'Jumlah']
        header_row = 8
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Map web items to desktop-like rows
        def split_category(name: str):
            if not name:
                return ('-', '-')
            # Try standard patterns
            if ' - ' in name:
                parts = name.split(' - ', 1)
                return (parts[0], parts[1])
            if name.startswith('Bed Frame'):
                return ('Bed Frame', '-')
            if name.startswith('Backdrop Panel'):
                # e.g., Backdrop Panel (Flat)
                sub = name.replace('Backdrop Panel', '').strip().strip('() ')
                return ('Backdrop Panel', sub or '-')
            return (name, '-')

        def normalize_dimensi(dim: str):
            if not dim:
                return ''
            d = dim.replace('cm', '').replace('CM', '').strip()
            d = d.replace('x', '×').replace(' X ', '×').replace(' x ', '×')
            return d

        for idx, it in enumerate(items_data, start=1):
            kategori, sub_item = split_category(it.get('nama_item', ''))
            # Prefer rich description, but fall back to finishing label if none
            if hasattr(MaterialDescriptions, 'get_description'):
                desc_try = MaterialDescriptions.get_description(it.get('finishing', ''))
                deskripsi = desc_try if desc_try and 'Tidak ada deskripsi' not in desc_try else f"Finishing {it.get('finishing','')}"
            else:
                deskripsi = it.get('finishing', '')
            dimensi = normalize_dimensi(it.get('dimensi', ''))
            volume = float(it.get('area', 0) or 0)
            harga_satuan = float(it.get('harga_satuan', 0) or 0)
            jumlah = float(it.get('jumlah', 0) or 0)

            row = [
                idx,
                kategori,
                sub_item,
                deskripsi,
                dimensi,
                'meter',
                volume,
                harga_satuan,
                jumlah
            ]

            row_num = header_row + idx
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = value
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                if col_idx in [8, 9]:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Number formatting
        for r in range(header_row + 1, header_row + 1 + len(items_data)):
            ws[f'G{r}'].number_format = '0.00'
            ws[f'H{r}'].number_format = '"Rp "#,##0'
            ws[f'I{r}'].number_format = '"Rp "#,##0'

        # Totals
        subtotal = sum(float(it.get('jumlah', 0) or 0) for it in items_data)
        discount_percent = float(pricing_info.get('discount_percent', 0) or 0)
        enable_tax = bool(pricing_info.get('enable_tax', False))
        tax_percent = float(pricing_info.get('tax_percent', 11) or 11)

        total_row = header_row + len(items_data) + 2
        ws.merge_cells(f'A{total_row}:H{total_row}')
        ws[f'A{total_row}'] = 'SUBTOTAL'
        ws[f'A{total_row}'].font = Font(bold=True, size=11)
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'I{total_row}'] = subtotal
        ws[f'I{total_row}'].font = Font(bold=True, size=11)
        ws[f'I{total_row}'].number_format = '"Rp "#,##0'
        ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')

        total_row += 1
        if discount_percent > 0:
            discount_amount = subtotal * (discount_percent / 100)
            ws.merge_cells(f'A{total_row}:H{total_row}')
            ws[f'A{total_row}'] = f'DISKON ({discount_percent}%)'
            ws[f'A{total_row}'].font = Font(size=10, color='e74c3c')
            ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'I{total_row}'] = -discount_amount
            ws[f'I{total_row}'].font = Font(size=10, color='e74c3c')
            ws[f'I{total_row}'].number_format = '"Rp "#,##0'
            ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
            total_row += 1
            after_discount = subtotal - discount_amount
        else:
            after_discount = subtotal

        if enable_tax:
            tax_amount = after_discount * (tax_percent / 100)
            ws.merge_cells(f'A{total_row}:H{total_row}')
            ws[f'A{total_row}'] = f'PAJAK/PPN ({tax_percent}%)'
            ws[f'A{total_row}'].font = Font(size=10)
            ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'I{total_row}'] = tax_amount
            ws[f'I{total_row}'].font = Font(size=10)
            ws[f'I{total_row}'].number_format = '"Rp "#,##0'
            ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
            total_row += 1
            grand_total = after_discount + tax_amount
        else:
            grand_total = after_discount

        ws.merge_cells(f'A{total_row}:H{total_row}')
        ws[f'A{total_row}'] = 'GRAND TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True, size=13, color='FFFFFF')
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{total_row}'].fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
        ws[f'I{total_row}'] = grand_total
        ws[f'I{total_row}'].font = Font(bold=True, size=13, color='FFFFFF')
        ws[f'I{total_row}'].number_format = '"Rp "#,##0'
        ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'I{total_row}'].fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')

        # Customer info block
        customer_name = customer_data.get('name', '')
        customer_address = customer_data.get('address', '')
        customer_phone = customer_data.get('phone', '')
        if customer_name or customer_address or customer_phone:
            total_row += 2
            ws.merge_cells(f'A{total_row}:I{total_row}')
            ws[f'A{total_row}'] = '👤 INFORMASI CUSTOMER'
            ws[f'A{total_row}'].font = Font(bold=True, size=11, color='2c3e50')
            ws[f'A{total_row}'].fill = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')
            ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
            total_row += 1
            if customer_name:
                ws.merge_cells(f'A{total_row}:I{total_row}')
                ws[f'A{total_row}'] = f'Nama: {customer_name}'
                ws[f'A{total_row}'].font = Font(size=10)
                ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
                total_row += 1
            if customer_address:
                ws.merge_cells(f'A{total_row}:I{total_row}')
                ws[f'A{total_row}'] = f'Alamat: {customer_address}'
                ws[f'A{total_row}'].font = Font(size=10)
                ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
                total_row += 1
            if customer_phone:
                ws.merge_cells(f'A{total_row}:I{total_row}')
                ws[f'A{total_row}'] = f'Telepon: {customer_phone}'
                ws[f'A{total_row}'].font = Font(size=10)
                ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 18

        # Filename like desktop: MPN<YEAR><ROMAN><SEQ>-<Customer>.xlsx
        now = datetime.now()
        year = now.strftime('%Y')
        roman_months = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII']
        month_roman = roman_months[now.month - 1]
        sequence = now.strftime('%d%H%M')
        customer_name_clean = (customer_name or 'NoName').strip()
        customer_name_clean = "".join(c for c in customer_name_clean if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f'MPN{year}{month_roman}{sequence}-{customer_name_clean}.xlsx'

        temp_dir = '/tmp' if os.environ.get('VERCEL') else tempfile.gettempdir()
        file_path = os.path.join(temp_dir, filename)
        wb.save(file_path)

        # WhatsApp message (simplified totals)
        item_list = "\n".join([
            f"{i+1}. {it.get('nama_item','')} - {it.get('finishing','')} - Rp{float(it.get('jumlah',0) or 0):,.0f}" for i, it in enumerate(items_data[:5])
        ])
        if len(items_data) > 5:
            item_list += f"\n... dan {len(items_data) - 5} item lainnya"

        whatsapp_message = (
            "🏢 *QUOTATION FURNITURE*\n"
            "PT. Enigma Prisma Delapan\n\n"
            f"👤 Customer: {customer_name}\n"
            f"📱 Telepon: {customer_phone}\n"
            f"📍 Alamat: {customer_address}\n"
            f"📅 Tanggal: {now.strftime('%d/%m/%Y')}\n\n"
            "📋 *ITEM LIST:*\n" + item_list + "\n\n"
            f"*GRAND TOTAL: Rp{subtotal:,.0f}*\n\n"
            "📎 File Excel quotation lengkap sudah dikirim!\n\n"
            "Terima kasih atas kepercayaan Anda! 🙏\n"
            "Jika ada pertanyaan, silakan hubungi kami.\n\n"
            "🌐 interiormapan.com\n📞 0821 1213 4258"
        )

        target_whatsapp = "6282112134258"
        whatsapp_result = send_whatsapp_message(target_whatsapp, whatsapp_message)

        file_url = f'/download_temp_excel/{filename}'
        return jsonify({
            'success': True,
            'filename': filename,
            'file_url': file_url,
            'whatsapp_number': target_whatsapp,
            'whatsapp_url': whatsapp_result.get('whatsapp_url', ''),
            'message': 'Excel berhasil dibuat dan pesan WhatsApp siap dikirim'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': f'Error: {str(e)}'}), 500

@app.route('/download_temp_excel/<filename>')
def download_temp_excel(filename):
    """Download temporary Excel file"""
    try:
        # Use proper temporary directory for serverless
        if os.environ.get('VERCEL'):
            temp_dir = '/tmp'
        else:
            temp_dir = tempfile.gettempdir()
        
        file_path = os.path.join(temp_dir, filename)
        
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True, download_name=filename,
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            return jsonify({'error': 'File not found'}), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# For Vercel deployment
app.debug = False

# Serverless compatibility
if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))