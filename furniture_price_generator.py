import sys
import os
import json
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QWidget, QLabel, QVBoxLayout, QPushButton, 
                           QLineEdit, QListWidget, QMessageBox, QComboBox, QHBoxLayout,
                           QFormLayout, QGroupBox, QScrollArea, QStackedWidget, QFileDialog,
                           QDialog, QDialogButtonBox, QSpinBox, QDoubleSpinBox, QShortcut)
from PyQt5.QtGui import QFont, QPixmap, QIcon, QDoubleValidator, QKeySequence
from PyQt5.QtCore import Qt
from info_combo import InfoComboBox
from material_descriptions import MaterialDescriptions

class FurniturePriceGenerator(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Furniture Price Generator by Mapan')
        self.setGeometry(100, 100, 1000, 750)  # Lebih besar untuk list area
        
        # Set window icon
        icon_path = os.path.join(os.path.dirname(__file__), 'logo.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        
        # Maximize window on startup
        self.showMaximized()
        
        self.is_dark_mode = False
        self.items_data = []
        
        # Customer & Pricing Info
        self.customer_name = ""
        self.customer_address = ""
        self.customer_phone = ""
        self.discount_percent = 0.0
        self.tax_percent = 11.0  # PPN 11%
        self.use_tax = True  # Flag untuk pakai pajak atau tidak
        
        self.apply_light_mode()
        self.initUI()
        self.setup_shortcuts()

    def initUI(self):
        main_layout = QVBoxLayout()
        
        # ========== HEADER SECTION ==========
        header_layout = QHBoxLayout()
        header_layout.setSpacing(20)
        
        # Logo di kiri
        logo_label = QLabel()
        
        # Coba load logo dari file
        import os
        logo_path = os.path.join(os.path.dirname(__file__), 'logo.png')
        
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            # Scale logo lebih kecil untuk header yang compact
            scaled_pixmap = pixmap.scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo_label.setPixmap(scaled_pixmap)
            logo_label.setAlignment(Qt.AlignCenter)
            logo_label.setFixedSize(80, 80)
        else:
            # Fallback ke text logo jika file tidak ada
            logo_label.setStyleSheet("""
                QLabel {
                    background-color: #c0392b;
                    color: white;
                    font-size: 24px;
                    font-weight: bold;
                    border-radius: 4px;
                    padding: 8px;
                }
            """)
            logo_label.setText("EPD")
            logo_label.setAlignment(Qt.AlignCenter)
            logo_label.setFixedSize(80, 80)
        
        header_layout.addWidget(logo_label)
        
        # Spacer
        header_layout.addStretch()
        
        # Info perusahaan di kanan (compact version)
        self.company_info = QLabel()
        self.company_info.setStyleSheet("""
            QLabel {
                color: #2c3e50;
                font-size: 9px;
                background-color: white;
                border-radius: 4px;
                padding: 6px;
            }
        """)
        self.company_info.setText(
            "<div style='text-align: right;'>"
            "<b style='font-size: 11px; color: #c0392b;'>PT. ENIGMA PRISMA DELAPAN</b><br>"
            "<span style='font-size: 8px;'>Jl. Raya H. Abdullah No.56, Pakulonan Barat<br>"
            "Tangerang - 0821 1213 4258</span>"
            "</div>"
        )
        self.company_info.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        header_layout.addWidget(self.company_info)
        
        main_layout.addLayout(header_layout)
        
        # Compact title and controls in one line
        title_layout = QHBoxLayout()
        
        # ========== TITLE ==========
        self.title = QLabel('Furniture Price Generator')
        self.title.setObjectName('TitleLabel')
        self.title.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        title_layout.addWidget(self.title)
        
        title_layout.addStretch()
        
        # Kategori dropdown inline
        title_layout.addWidget(QLabel('Kategori:'))
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItems([
            'Kitchen Set', 'Wardrobe', 'Bed', 'Backdrop Panel',
            'Credenza', 'Multi Cabinet', 'Custom Furniture'
        ])
        self.kategori_combo.currentIndexChanged.connect(self.update_form)
        self.kategori_combo.setMinimumWidth(150)
        title_layout.addWidget(self.kategori_combo)
        
        # Dark Mode Toggle Button inline
        self.theme_toggle_btn = QPushButton('🌙')
        self.theme_toggle_btn.setObjectName('ThemeToggle')
        self.theme_toggle_btn.clicked.connect(self.toggle_dark_mode)
        self.theme_toggle_btn.setFixedWidth(40)
        self.theme_toggle_btn.setToolTip('Toggle Dark Mode')
        title_layout.addWidget(self.theme_toggle_btn)
        
        main_layout.addLayout(title_layout)
        
        # Separator line
        self.separator = QLabel()
        self.separator.setStyleSheet("background-color: #c0392b; max-height: 2px;")
        self.separator.setFixedHeight(2)
        main_layout.addWidget(self.separator)
        main_layout.addSpacing(5)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)

        self.stacked_widget = QStackedWidget()
        
        self.kitchen_widget = self.create_kitchen_form()
        self.wardrobe_widget = self.create_wardrobe_form()
        self.bed_widget = self.create_bed_form()
        self.backdrop_widget = self.create_backdrop_form()
        self.credenza_widget = self.create_credenza_form()
        self.multi_widget = self.create_multi_cabinet_form()
        self.custom_widget = self.create_custom_form()
        
        self.stacked_widget.addWidget(self.kitchen_widget)
        self.stacked_widget.addWidget(self.wardrobe_widget)
        self.stacked_widget.addWidget(self.bed_widget)
        self.stacked_widget.addWidget(self.backdrop_widget)
        self.stacked_widget.addWidget(self.credenza_widget)
        self.stacked_widget.addWidget(self.multi_widget)
        self.stacked_widget.addWidget(self.custom_widget)
        
        scroll_layout.addWidget(self.stacked_widget)
        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll)

        self.add_button = QPushButton('➕ Tambah Item')
        self.add_button.clicked.connect(self.add_item)
        main_layout.addWidget(self.add_button)

        self.list_widget = QListWidget()
        self.list_widget.setMinimumHeight(250)  # Set minimum height untuk list
        main_layout.addWidget(self.list_widget)
        
        # Running Total Display
        self.total_display = QLabel()
        self.total_display.setStyleSheet("""
            QLabel {
                background-color: #e8f5e9;
                border: 2px solid #4caf50;
                border-radius: 6px;
                padding: 12px;
                font-size: 16px;
                font-weight: bold;
                color: #2e7d32;
            }
        """)
        self.update_total_display()
        main_layout.addWidget(self.total_display)

        # Item Management Buttons
        item_btn_layout = QHBoxLayout()
        
        self.delete_button = QPushButton('🗑️ Hapus Item Terpilih')
        self.delete_button.clicked.connect(self.delete_selected_item)
        item_btn_layout.addWidget(self.delete_button)
        
        self.clear_button = QPushButton('🗑️ Hapus Semua Item')
        self.clear_button.clicked.connect(self.clear_all_items)
        item_btn_layout.addWidget(self.clear_button)
        
        main_layout.addLayout(item_btn_layout)

        # Main Action Buttons
        button_layout = QHBoxLayout()
        
        self.customer_button = QPushButton('👤 Info Customer')
        self.customer_button.clicked.connect(self.show_customer_dialog)
        button_layout.addWidget(self.customer_button)
        
        self.discount_button = QPushButton('💰 Diskon & Pajak')
        self.discount_button.clicked.connect(self.show_discount_dialog)
        button_layout.addWidget(self.discount_button)
        
        self.generate_button = QPushButton('📊 Hitung Total')
        self.generate_button.clicked.connect(self.generate_total)
        button_layout.addWidget(self.generate_button)
        
        main_layout.addLayout(button_layout)
        
        # File Operations
        file_layout = QHBoxLayout()
        
        self.save_button = QPushButton('💾 Save Project')
        self.save_button.clicked.connect(self.save_project)
        file_layout.addWidget(self.save_button)
        
        self.load_button = QPushButton('📂 Load Project')
        self.load_button.clicked.connect(self.load_project)
        file_layout.addWidget(self.load_button)
        
        self.export_button = QPushButton('📄 Export ke Excel')
        self.export_button.clicked.connect(self.export_to_excel)
        file_layout.addWidget(self.export_button)
        
        main_layout.addLayout(file_layout)
        self.setLayout(main_layout)
        self.update_form()
    
    def create_validated_lineedit(self, placeholder=""):
        """Create a LineEdit with number validation"""
        line_edit = QLineEdit()
        line_edit.setPlaceholderText(placeholder)
        # Allow only positive numbers with decimals
        validator = QDoubleValidator(0.0, 999999.99, 2)
        validator.setNotation(QDoubleValidator.StandardNotation)
        line_edit.setValidator(validator)
        return line_edit

    def create_kitchen_form(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Ilustrasi gambar pengukuran
        ilustrasi_label = QLabel()
        ilustrasi_label.setAlignment(Qt.AlignCenter)
        
        import os
        ilustrasi_path = os.path.join(os.path.dirname(__file__), 'illustration_kitchen.png')
        
        if os.path.exists(ilustrasi_path):
            pixmap = QPixmap(ilustrasi_path)
            scaled_pixmap = pixmap.scaled(400, 300, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            ilustrasi_label.setPixmap(scaled_pixmap)
        else:
            # Fallback ke text jika gambar tidak ada
            ilustrasi_label.setStyleSheet("""
                QLabel {
                    background-color: #fff9e6;
                    border: 2px dashed #f39c12;
                    border-radius: 6px;
                    padding: 10px;
                    color: #856404;
                    font-size: 11px;
                }
            """)
            ilustrasi_label.setText(
                "📏 <b>PANDUAN PENGUKURAN KITCHEN SET:</b><br>"
                "Gambar ilustrasi tidak ditemukan. Jalankan: python create_measurement_illustrations.py"
            )
            ilustrasi_label.setWordWrap(True)
        
        layout.addWidget(ilustrasi_label)
        
        kb_group = QGroupBox('Kabinet Bawah')
        kb_layout = QFormLayout()
        self.kb_panjang = self.create_validated_lineedit("Contoh: 300")
        self.kb_tinggi = self.create_validated_lineedit("Contoh: 85")
        self.kb_tebal = self.create_validated_lineedit("Contoh: 60")
        self.kb_finishing = InfoComboBox(['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'])
        
        # Label keterangan kelas HPL
        self.kb_hpl_info = QLabel()
        self.kb_hpl_info.setWordWrap(True)
        self.kb_hpl_info.setStyleSheet("""
            QLabel {
                background-color: #e8f5e9;
                border: 1px solid #4caf50;
                border-radius: 4px;
                padding: 8px;
                color: #2e7d32;
                font-size: 10px;
            }
        """)
        self.kb_hpl_info.setText("""💡 Perbedaan Kelas HPL:
• Tacosheet (Rp 2,3jt/m²): Lembaran HPL tipis, budget friendly
• HPL Low (Rp 2,3jt/m²): Luar HPL, dalam melamninto, motif terbatas
• HPL Mid (Rp 2,8jt/m²): Luar HPL, dalam melamninto, motif standar
• HPL High (Rp 3,2jt/m²): Luar+dalam full HPL, semua motif tersedia
• Duco (Rp 5,5jt/m²): Full cat duco custom warna
• Kombinasi (Rp 4,7jt/m²): Body HPL + pintu duco""")
        
        kb_layout.addRow('Panjang (cm):', self.kb_panjang)
        kb_layout.addRow('Tinggi (cm):', self.kb_tinggi)
        kb_layout.addRow('Tebal (cm):', self.kb_tebal)
        kb_layout.addRow('Finishing:', self.kb_finishing)
        kb_layout.addRow('', self.kb_hpl_info)
        kb_group.setLayout(kb_layout)
        layout.addWidget(kb_group)
        
        ka_group = QGroupBox('Kabinet Atas')
        ka_layout = QFormLayout()
        self.ka_panjang = QLineEdit()
        self.ka_tinggi = QLineEdit()
        self.ka_tebal = QLineEdit()
        self.ka_finishing = InfoComboBox(['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'])
        
        # Label keterangan kelas HPL
        self.ka_hpl_info = QLabel()
        self.ka_hpl_info.setWordWrap(True)
        self.ka_hpl_info.setStyleSheet("""
            QLabel {
                background-color: #e8f5e9;
                border: 1px solid #4caf50;
                border-radius: 4px;
                padding: 8px;
                color: #2e7d32;
                font-size: 10px;
            }
        """)
        self.ka_hpl_info.setText("""💡 Perbedaan Kelas HPL:
• Tacosheet (Rp 2,3jt/m²): Lembaran HPL tipis, budget friendly
• HPL Low (Rp 2,3jt/m²): Luar HPL, dalam melamninto, motif terbatas
• HPL Mid (Rp 2,8jt/m²): Luar HPL, dalam melamninto, motif standar
• HPL High (Rp 3,2jt/m²): Luar+dalam full HPL, semua motif tersedia
• Duco (Rp 5,5jt/m²): Full cat duco custom warna
• Kombinasi (Rp 4,7jt/m²): Body HPL + pintu duco""")
        
        ka_layout.addRow('Panjang (cm):', self.ka_panjang)
        ka_layout.addRow('Tinggi (cm):', self.ka_tinggi)
        ka_layout.addRow('Tebal (cm):', self.ka_tebal)
        ka_layout.addRow('Finishing:', self.ka_finishing)
        ka_layout.addRow('', self.ka_hpl_info)
        ka_group.setLayout(ka_layout)
        layout.addWidget(ka_group)
        
        tt_group = QGroupBox('Top Table')
        tt_layout = QFormLayout()
        self.tt_panjang = QLineEdit()
        self.tt_lebar = QLineEdit()
        self.tt_material = InfoComboBox(['Solid Surface', 'Granit Alam', 'Marmer'])
        tt_layout.addRow('Panjang (cm):', self.tt_panjang)
        tt_layout.addRow('Lebar (cm):', self.tt_lebar)
        tt_layout.addRow('Material:', self.tt_material)
        tt_group.setLayout(tt_layout)
        layout.addWidget(tt_group)
        
        bs_group = QGroupBox('Backsplash')
        bs_layout = QFormLayout()
        self.bs_panjang = QLineEdit()
        self.bs_tinggi = QLineEdit()
        self.bs_material = InfoComboBox([
            'Solid Surface', 'Granit Alam', 'Marmer',
            'Mirror Clear', 'Bronze Mirror', 'Keramik'
        ])
        bs_layout.addRow('Panjang (cm):', self.bs_panjang)
        bs_layout.addRow('Tinggi (cm):', self.bs_tinggi)
        bs_layout.addRow('Material:', self.bs_material)
        bs_group.setLayout(bs_layout)
        layout.addWidget(bs_group)
        
        widget.setLayout(layout)
        return widget

    def create_wardrobe_form(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Ilustrasi gambar pengukuran
        ilustrasi_label = QLabel()
        ilustrasi_label.setAlignment(Qt.AlignCenter)
        
        import os
        ilustrasi_path = os.path.join(os.path.dirname(__file__), 'illustration_wardrobe.png')
        
        if os.path.exists(ilustrasi_path):
            pixmap = QPixmap(ilustrasi_path)
            scaled_pixmap = pixmap.scaled(400, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            ilustrasi_label.setPixmap(scaled_pixmap)
        else:
            ilustrasi_label.setText("📏 Gambar ilustrasi tidak ditemukan")
        
        layout.addWidget(ilustrasi_label)
        
        group = QGroupBox('Wardrobe')
        form = QFormLayout()
        self.wr_panjang = QLineEdit()
        self.wr_tinggi = QLineEdit()
        self.wr_tebal = QLineEdit()
        self.wr_finishing = InfoComboBox(['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'])
        
        # Label keterangan kelas HPL
        self.wr_hpl_info = QLabel()
        self.wr_hpl_info.setWordWrap(True)
        self.wr_hpl_info.setStyleSheet("""
            QLabel {
                background-color: #e8f5e9;
                border: 1px solid #4caf50;
                border-radius: 4px;
                padding: 8px;
                color: #2e7d32;
                font-size: 10px;
            }
        """)
        self.wr_hpl_info.setText("""💡 Perbedaan Kelas HPL:
• Tacosheet (Rp 2,3jt/m²): Lembaran HPL tipis, budget friendly
• HPL Low (Rp 2,3jt/m²): Luar HPL, dalam melamninto, motif terbatas
• HPL Mid (Rp 2,8jt/m²): Luar HPL, dalam melamninto, motif standar
• HPL High (Rp 3,2jt/m²): Luar+dalam full HPL, semua motif tersedia
• Duco (Rp 5,5jt/m²): Full cat duco custom warna
• Kombinasi (Rp 4,7jt/m²): Body HPL + pintu duco""")
        
        form.addRow('Panjang (cm):', self.wr_panjang)
        form.addRow('Tinggi (cm):', self.wr_tinggi)
        form.addRow('Tebal (cm):', self.wr_tebal)
        form.addRow('Finishing:', self.wr_finishing)
        form.addRow('', self.wr_hpl_info)
        group.setLayout(form)
        layout.addWidget(group)
        
        widget.setLayout(layout)
        return widget

    def create_bed_form(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Ilustrasi gambar pengukuran
        ilustrasi_label = QLabel()
        ilustrasi_label.setAlignment(Qt.AlignCenter)
        
        import os
        ilustrasi_path = os.path.join(os.path.dirname(__file__), 'illustration_bed.png')
        
        if os.path.exists(ilustrasi_path):
            pixmap = QPixmap(ilustrasi_path)
            scaled_pixmap = pixmap.scaled(500, 350, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            ilustrasi_label.setPixmap(scaled_pixmap)
        else:
            ilustrasi_label.setText("📏 Gambar ilustrasi tidak ditemukan")
        
        layout.addWidget(ilustrasi_label)
        
        frame_group = QGroupBox('Bed Frame')
        frame_layout = QFormLayout()
        self.bed_panjang = QLineEdit()
        self.bed_lebar = QLineEdit()
        self.bed_tinggi = QLineEdit()
        self.bed_finishing = InfoComboBox(['HPL', 'Duco'])
        
        frame_layout.addRow('Panjang (cm):', self.bed_panjang)
        frame_layout.addRow('Lebar (cm):', self.bed_lebar)
        frame_layout.addRow('Tinggi (cm):', self.bed_tinggi)
        frame_layout.addRow('Finishing:', self.bed_finishing)
        frame_group.setLayout(frame_layout)
        layout.addWidget(frame_group)
        
        hb_group = QGroupBox('Headboard')
        hb_layout = QFormLayout()
        self.hb_panjang = QLineEdit()
        self.hb_tinggi = QLineEdit()
        self.hb_material = InfoComboBox(['Synthetic Leather', 'Fabric'])
        hb_layout.addRow('Panjang (cm):', self.hb_panjang)
        hb_layout.addRow('Tinggi (cm):', self.hb_tinggi)
        hb_layout.addRow('Material:', self.hb_material)
        hb_group.setLayout(hb_layout)
        layout.addWidget(hb_group)
        
        widget.setLayout(layout)
        return widget

    def create_backdrop_form(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Ilustrasi gambar pengukuran
        ilustrasi_label = QLabel()
        ilustrasi_label.setAlignment(Qt.AlignCenter)
        
        import os
        ilustrasi_path = os.path.join(os.path.dirname(__file__), 'illustration_backdrop.png')
        
        if os.path.exists(ilustrasi_path):
            pixmap = QPixmap(ilustrasi_path)
            scaled_pixmap = pixmap.scaled(400, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            ilustrasi_label.setPixmap(scaled_pixmap)
        else:
            ilustrasi_label.setText("📏 Gambar ilustrasi tidak ditemukan")
        
        layout.addWidget(ilustrasi_label)
        
        group = QGroupBox('Backdrop Panel')
        form = QFormLayout()
        self.bd_panjang = QLineEdit()
        self.bd_tinggi = QLineEdit()
        self.bd_type = QComboBox()
        self.bd_type.addItems(['PVC', 'Flat', 'Tebal'])
        self.bd_type.currentIndexChanged.connect(self.update_backdrop_finishing)
        
        # Label keterangan type backdrop
        self.bd_type_info = QLabel()
        self.bd_type_info.setWordWrap(True)
        self.bd_type_info.setStyleSheet("""
            QLabel {
                background-color: #e8f4f8;
                border: 1px solid #4a90e2;
                border-radius: 4px;
                padding: 8px;
                color: #2c5aa0;
                font-size: 11px;
            }
        """)
        
        self.bd_finishing = InfoComboBox(['HPL', 'Duco', 'Kombinasi'])
        
        form.addRow('Panjang (cm):', self.bd_panjang)
        form.addRow('Tinggi (cm):', self.bd_tinggi)
        form.addRow('Type:', self.bd_type)
        form.addRow('', self.bd_type_info)
        form.addRow('Finishing:', self.bd_finishing)
        group.setLayout(form)
        layout.addWidget(group)
        
        # Connect signal dan set keterangan awal
        self.bd_type.currentTextChanged.connect(self.update_backdrop_type_info)
        self.update_backdrop_type_info()
        
        widget.setLayout(layout)
        return widget

    def create_credenza_form(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Ilustrasi gambar pengukuran
        ilustrasi_label = QLabel()
        ilustrasi_label.setAlignment(Qt.AlignCenter)
        
        import os
        ilustrasi_path = os.path.join(os.path.dirname(__file__), 'illustration_credenza.png')
        
        if os.path.exists(ilustrasi_path):
            pixmap = QPixmap(ilustrasi_path)
            scaled_pixmap = pixmap.scaled(450, 300, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            ilustrasi_label.setPixmap(scaled_pixmap)
        else:
            ilustrasi_label.setText("📏 Gambar ilustrasi tidak ditemukan")
        
        layout.addWidget(ilustrasi_label)
        
        group = QGroupBox('Credenza')
        form = QFormLayout()
        self.cr_panjang = QLineEdit()
        self.cr_tinggi = QLineEdit()
        self.cr_tebal = QLineEdit()
        self.cr_finishing = InfoComboBox(['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'])
        
        # Label keterangan kelas HPL
        self.cr_hpl_info = QLabel()
        self.cr_hpl_info.setWordWrap(True)
        self.cr_hpl_info.setStyleSheet("""
            QLabel {
                background-color: #e8f5e9;
                border: 1px solid #4caf50;
                border-radius: 4px;
                padding: 8px;
                color: #2e7d32;
                font-size: 10px;
            }
        """)
        self.cr_hpl_info.setText("""💡 Perbedaan Kelas HPL:
• Tacosheet (Rp 2,3jt/m²): Lembaran HPL tipis, budget friendly
• HPL Low (Rp 2,3jt/m²): Luar HPL, dalam melamninto, motif terbatas
• HPL Mid (Rp 2,8jt/m²): Luar HPL, dalam melamninto, motif standar
• HPL High (Rp 3,2jt/m²): Luar+dalam full HPL, semua motif tersedia
• Duco (Rp 5,5jt/m²): Full cat duco custom warna
• Kombinasi (Rp 4,7jt/m²): Body HPL + pintu duco""")
        
        form.addRow('Panjang (cm):', self.cr_panjang)
        form.addRow('Tinggi (cm):', self.cr_tinggi)
        form.addRow('Tebal (cm):', self.cr_tebal)
        form.addRow('Finishing:', self.cr_finishing)
        form.addRow('', self.cr_hpl_info)
        group.setLayout(form)
        layout.addWidget(group)
        
        widget.setLayout(layout)
        return widget

    def create_multi_cabinet_form(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Ilustrasi gambar pengukuran (pakai gambar kitchen karena serupa)
        ilustrasi_label = QLabel()
        ilustrasi_label.setAlignment(Qt.AlignCenter)
        
        import os
        ilustrasi_path = os.path.join(os.path.dirname(__file__), 'illustration_kitchen.png')
        
        if os.path.exists(ilustrasi_path):
            pixmap = QPixmap(ilustrasi_path)
            scaled_pixmap = pixmap.scaled(400, 300, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            ilustrasi_label.setPixmap(scaled_pixmap)
        else:
            ilustrasi_label.setText("📏 Gambar ilustrasi tidak ditemukan")
        
        layout.addWidget(ilustrasi_label)
        
        group = QGroupBox('Multi Cabinet')
        form = QFormLayout()
        self.mc_panjang = QLineEdit()
        self.mc_tinggi = QLineEdit()
        self.mc_tebal = QLineEdit()
        self.mc_finishing = InfoComboBox(['Tacosheet', 'HPL Low', 'HPL Mid', 'HPL High', 'Duco', 'Kombinasi'])
        
        # Label keterangan kelas HPL
        self.mc_hpl_info = QLabel()
        self.mc_hpl_info.setWordWrap(True)
        self.mc_hpl_info.setStyleSheet("""
            QLabel {
                background-color: #e8f5e9;
                border: 1px solid #4caf50;
                border-radius: 4px;
                padding: 8px;
                color: #2e7d32;
                font-size: 10px;
            }
        """)
        self.mc_hpl_info.setText("""💡 Perbedaan Kelas HPL:
• Tacosheet (Rp 2,3jt/m²): Lembaran HPL tipis, budget friendly
• HPL Low (Rp 2,3jt/m²): Luar HPL, dalam melamninto, motif terbatas
• HPL Mid (Rp 2,8jt/m²): Luar HPL, dalam melamninto, motif standar
• HPL High (Rp 3,2jt/m²): Luar+dalam full HPL, semua motif tersedia
• Duco (Rp 5,5jt/m²): Full cat duco custom warna
• Kombinasi (Rp 4,7jt/m²): Body HPL + pintu duco""")
        
        form.addRow('Panjang (cm):', self.mc_panjang)
        form.addRow('Tinggi (cm):', self.mc_tinggi)
        form.addRow('Tebal (cm):', self.mc_tebal)
        form.addRow('Finishing:', self.mc_finishing)
        form.addRow('', self.mc_hpl_info)
        group.setLayout(form)
        layout.addWidget(group)
        
        widget.setLayout(layout)
        return widget

    def create_custom_form(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        # Ilustrasi gambar pengukuran
        ilustrasi_label = QLabel()
        ilustrasi_label.setAlignment(Qt.AlignCenter)
        
        import os
        ilustrasi_path = os.path.join(os.path.dirname(__file__), 'illustration_custom.png')
        
        if os.path.exists(ilustrasi_path):
            pixmap = QPixmap(ilustrasi_path)
            scaled_pixmap = pixmap.scaled(400, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            ilustrasi_label.setPixmap(scaled_pixmap)
        else:
            ilustrasi_label.setText("📏 Gambar ilustrasi tidak ditemukan")
        
        layout.addWidget(ilustrasi_label)
        
        group = QGroupBox('Custom Furniture')
        form = QFormLayout()
        self.custom_nama = QLineEdit()
        self.custom_panjang = QLineEdit()
        self.custom_lebar = QLineEdit()
        self.custom_tinggi = QLineEdit()
        self.custom_method = QComboBox()
        self.custom_method.addItems([
            'Meter Lari', 'Meter Persegi', 'Side Area X', 'Side Area Y', 'Side Area Z'
        ])
        
        # Label keterangan metode
        self.custom_method_info = QLabel()
        self.custom_method_info.setWordWrap(True)
        self.custom_method_info.setStyleSheet("""
            QLabel {
                background-color: #fff3cd;
                border: 1px solid #ffc107;
                border-radius: 4px;
                padding: 8px;
                color: #856404;
                font-size: 11px;
            }
        """)
        
        self.custom_unit_price = QLineEdit()
        form.addRow('Nama Item:', self.custom_nama)
        form.addRow('Panjang (cm):', self.custom_panjang)
        form.addRow('Lebar (cm):', self.custom_lebar)
        form.addRow('Tinggi (cm):', self.custom_tinggi)
        form.addRow('Metode Hitung:', self.custom_method)
        form.addRow('', self.custom_method_info)
        form.addRow('Harga Satuan (Rp):', self.custom_unit_price)
        group.setLayout(form)
        layout.addWidget(group)
        
        # Connect signal untuk update keterangan
        self.custom_method.currentTextChanged.connect(self.update_custom_method_info)
        
        # Set keterangan awal
        self.update_custom_method_info()
        
        widget.setLayout(layout)
        return widget

    def update_form(self):
        self.stacked_widget.setCurrentIndex(self.kategori_combo.currentIndex())

    def update_custom_method_info(self):
        """Update keterangan metode perhitungan custom"""
        method = self.custom_method.currentText()
        
        info_text = {
            'Meter Lari': '📏 <b>Meter Lari</b><br>Rumus: Panjang ÷ 100<br>Contoh: Panjang 300cm = 3.0 meter<br>Digunakan untuk: Kitchen set, backsplash, list, molding',
            
            'Meter Persegi': '📐 <b>Meter Persegi</b><br>Rumus: (Panjang × Lebar) ÷ 10,000<br>Contoh: 200×100cm = 2.0 m²<br>Digunakan untuk: Panel datar, top table, partisi',
            
            'Side Area X': '🔴 <b>Side Area X (Sisi Samping)</b><br>Rumus: (Lebar × Tinggi) ÷ 10,000<br>Contoh: L=60cm, T=200cm = 1.2 m²<br>Digunakan untuk: Sisi kiri/kanan lemari',
            
            'Side Area Y': '🟢 <b>Side Area Y (Sisi Depan/Belakang)</b><br>Rumus: (Panjang × Tinggi) ÷ 10,000<br>Contoh: P=300cm, T=85cm = 2.55 m²<br>Digunakan untuk: Kabinet, backdrop, wardrobe (PALING UMUM)',
            
            'Side Area Z': '🔵 <b>Side Area Z (Sisi Atas/Bawah)</b><br>Rumus: (Panjang × Lebar) ÷ 10,000<br>Contoh: P=150cm, L=80cm = 1.2 m²<br>Digunakan untuk: Permukaan atas meja, rak horizontal'
        }
        
        self.custom_method_info.setText(info_text.get(method, ''))

    def update_backdrop_type_info(self):
        """Update keterangan type backdrop"""
        bd_type = self.bd_type.currentText()
        
        info_text = {
            'PVC': '🎨 <b>PVC Backdrop</b><br>Material: PVC Panel Motif<br>Harga: Rp 1.500.000/m²<br>Finishing: Tidak perlu (sudah jadi)',
            
            'Flat': '📋 <b>Flat Panel (Tipis)</b><br>Ketebalan: < 10cm<br>Multiplier: 1.0x<br>Finishing: Perlu (HPL/Duco/Kombinasi)',
            
            'Tebal': '📦 <b>Panel Tebal (3D)</b><br>Ketebalan: ≥ 10cm<br>Multiplier: 1.3x (lebih mahal)<br>Finishing: Perlu (HPL/Duco/Kombinasi)'
        }
        
        self.bd_type_info.setText(info_text.get(bd_type, ''))

    def update_backdrop_finishing(self):
        if self.bd_type.currentText() == 'PVC':
            self.bd_finishing.setEnabled(False)
        else:
            self.bd_finishing.setEnabled(True)

    def get_finishing_price(self, finishing):
        prices = {
            'Tacosheet': 2300000,
            'HPL Low': 2300000,
            'HPL Mid': 2800000,
            'HPL High': 3200000,
            'Duco': 5500000,
            'Kombinasi': 4700000
        }
        return prices.get(finishing, 0)

    def calculate_cabinet_price(self, panjang, tinggi, tebal, finishing):
        panjang = max(panjang, 100)
        tinggi = max(tinggi, 100)
        area = (panjang / 100) * (tinggi / 100)
        base_price = self.get_finishing_price(finishing)
        multiplier = 1.2 if tebal > 60 else 1.0
        total = area * base_price * multiplier
        return total, area, base_price * multiplier

    def add_item(self):
        kategori = self.kategori_combo.currentText()
        try:
            if kategori == 'Kitchen Set':
                self.add_kitchen_item()
            elif kategori == 'Wardrobe':
                self.add_wardrobe_item()
            elif kategori == 'Bed':
                self.add_bed_item()
            elif kategori == 'Backdrop Panel':
                self.add_backdrop_item()
            elif kategori == 'Credenza':
                self.add_credenza_item()
            elif kategori == 'Multi Cabinet':
                self.add_multi_cabinet_item()
            elif kategori == 'Custom Furniture':
                self.add_custom_item()
            
            # Update total display after adding item
            self.update_total_display()
        except ValueError as e:
            QMessageBox.warning(self, 'Error', f'Input tidak valid: {str(e)}')

    def add_kitchen_item(self):
        # Kabinet Bawah
        if all([self.kb_panjang.text(), self.kb_tinggi.text(), self.kb_tebal.text()]):
            finishing = self.kb_finishing.currentText()
            panjang = float(self.kb_panjang.text())
            tinggi = float(self.kb_tinggi.text())
            tebal = float(self.kb_tebal.text())
            
            total, area, unit_price = self.calculate_cabinet_price(panjang, tinggi, tebal, finishing)
            
            self.list_widget.addItem(
                f"Kitchen Set - Kabinet Bawah | {finishing} | "
                f"{max(panjang,100)}x{max(tinggi,100)}x{tebal}cm | Rp{total:,.0f}"
            )
            
            self.items_data.append({
                'nama_item': 'Kitchen Set',
                'sub_item': 'Kabinet Bawah',
                'deskripsi': MaterialDescriptions.get_description(finishing),
                'dimensi': f"{max(panjang,100)} x {max(tinggi,100)} x {tebal}",
                'satuan_dimensi': 'cm',
                'total_volume': area,
                'harga_dasar': unit_price,
                'jumlah': total
            })
            
            self.kb_panjang.clear()
            self.kb_tinggi.clear()
            self.kb_tebal.clear()
        
        # Kabinet Atas
        if all([self.ka_panjang.text(), self.ka_tinggi.text(), self.ka_tebal.text()]):
            finishing = self.ka_finishing.currentText()
            panjang = float(self.ka_panjang.text())
            tinggi = float(self.ka_tinggi.text())
            tebal = float(self.ka_tebal.text())
            
            total, area, unit_price = self.calculate_cabinet_price(panjang, tinggi, tebal, finishing)
            
            self.list_widget.addItem(
                f"Kitchen Set - Kabinet Atas | {finishing} | "
                f"{max(panjang,100)}x{max(tinggi,100)}x{tebal}cm | Rp{total:,.0f}"
            )
            
            self.items_data.append({
                'nama_item': 'Kitchen Set',
                'sub_item': 'Kabinet Atas',
                'deskripsi': MaterialDescriptions.get_description(finishing),
                'dimensi': f"{max(panjang,100)} x {max(tinggi,100)} x {tebal}",
                'satuan_dimensi': 'cm',
                'total_volume': area,
                'harga_dasar': unit_price,
                'jumlah': total
            })
            
            self.ka_panjang.clear()
            self.ka_tinggi.clear()
            self.ka_tebal.clear()
        
        # Top Table
        if all([self.tt_panjang.text(), self.tt_lebar.text()]):
            panjang = float(self.tt_panjang.text())
            lebar = float(self.tt_lebar.text())
            material = self.tt_material.currentText()
            
            material_prices = {
                'Solid Surface': 2000000,
                'Granit Alam': 2500000,
                'Marmer': 3500000
            }
            
            unit_price = material_prices.get(material, 0)
            meter_lari = panjang / 100
            total = unit_price * meter_lari
            
            self.list_widget.addItem(
                f"Kitchen Set - Top Table | {material} | {panjang}cm x {lebar}cm | Rp{total:,.0f}"
            )
            
            self.items_data.append({
                'nama_item': 'Kitchen Set',
                'sub_item': 'Top Table',
                'deskripsi': MaterialDescriptions.get_description(material),
                'dimensi': f"{panjang} x {lebar}",
                'satuan_dimensi': 'cm',
                'total_volume': meter_lari,
                'harga_dasar': unit_price,
                'jumlah': total
            })
            
            self.tt_panjang.clear()
            self.tt_lebar.clear()
        
        # Backsplash
        if all([self.bs_panjang.text(), self.bs_tinggi.text()]):
            panjang = float(self.bs_panjang.text())
            tinggi = float(self.bs_tinggi.text())
            material = self.bs_material.currentText()
            
            material_prices = {
                'Solid Surface': 2000000,
                'Granit Alam': 2500000,
                'Marmer': 3500000,
                'Mirror Clear': 1500000,
                'Bronze Mirror': 2000000,
                'Keramik': 850000
            }
            
            unit_price = material_prices.get(material, 0)
            meter_lari = panjang / 100
            total = unit_price * meter_lari
            
            self.list_widget.addItem(
                f"Kitchen Set - Backsplash | {material} | {panjang}cm x {tinggi}cm | Rp{total:,.0f}"
            )
            
            self.items_data.append({
                'nama_item': 'Kitchen Set',
                'sub_item': 'Backsplash',
                'deskripsi': MaterialDescriptions.get_description(material),
                'dimensi': f"{panjang} x {tinggi}",
                'satuan_dimensi': 'cm',
                'total_volume': meter_lari,
                'harga_dasar': unit_price,
                'jumlah': total
            })
            
            self.bs_panjang.clear()
            self.bs_tinggi.clear()

    def add_wardrobe_item(self):
        if not all([self.wr_panjang.text(), self.wr_tinggi.text(), self.wr_tebal.text()]):
            raise ValueError("Semua field harus diisi!")
        
        panjang = float(self.wr_panjang.text())
        tinggi = float(self.wr_tinggi.text())
        tebal = float(self.wr_tebal.text())
        finishing = self.wr_finishing.currentText()
        
        total, area, unit_price = self.calculate_cabinet_price(panjang, tinggi, tebal, finishing)
        
        self.list_widget.addItem(
            f"Wardrobe | {finishing} | {max(panjang,100)}x{max(tinggi,100)}x{tebal}cm | Rp{total:,.0f}"
        )
        
        self.items_data.append({
            'nama_item': 'Wardrobe',
            'sub_item': '-',
            'deskripsi': MaterialDescriptions.get_description(finishing),
            'dimensi': f"{max(panjang,100)} x {max(tinggi,100)} x {tebal}",
            'satuan_dimensi': 'cm',
            'total_volume': area,
            'harga_dasar': unit_price,
            'jumlah': total
        })
        
        self.wr_panjang.clear()
        self.wr_tinggi.clear()
        self.wr_tebal.clear()

    def add_bed_item(self):
        # Bed Frame
        if all([self.bed_panjang.text(), self.bed_lebar.text(), self.bed_tinggi.text()]):
            panjang = float(self.bed_panjang.text())
            lebar = float(self.bed_lebar.text())
            tinggi = float(self.bed_tinggi.text())
            finishing = self.bed_finishing.currentText()
            
            area = (panjang / 100) * (lebar / 100)
            # Bed menggunakan harga lama
            bed_prices = {'HPL': 3000000, 'Duco': 5500000}
            unit_price = bed_prices.get(finishing, 0)
            total = area * unit_price
            
            self.list_widget.addItem(
                f"Bed - Frame | {finishing} | {panjang}x{lebar}x{tinggi}cm | Rp{total:,.0f}"
            )
            
            self.items_data.append({
                'nama_item': 'Bed',
                'sub_item': 'Frame',
                'deskripsi': MaterialDescriptions.get_description(finishing),
                'dimensi': f"{panjang} x {lebar} x {tinggi}",
                'satuan_dimensi': 'cm',
                'total_volume': area,
                'harga_dasar': unit_price,
                'jumlah': total
            })
            
            self.bed_panjang.clear()
            self.bed_lebar.clear()
            self.bed_tinggi.clear()
        
        # Headboard
        if all([self.hb_panjang.text(), self.hb_tinggi.text()]):
            panjang = float(self.hb_panjang.text())
            tinggi = float(self.hb_tinggi.text())
            material = self.hb_material.currentText()
            
            area = (panjang / 100) * (tinggi / 100)
            unit_price = 1900000
            total = area * unit_price
            
            self.list_widget.addItem(
                f"Bed - Headboard | {material} | {panjang}x{tinggi}cm | Rp{total:,.0f}"
            )
            
            self.items_data.append({
                'nama_item': 'Bed',
                'sub_item': 'Headboard',
                'deskripsi': MaterialDescriptions.get_description(material),
                'dimensi': f"{panjang} x {tinggi}",
                'satuan_dimensi': 'cm',
                'total_volume': area,
                'harga_dasar': unit_price,
                'jumlah': total
            })
            
            self.hb_panjang.clear()
            self.hb_tinggi.clear()

    def add_backdrop_item(self):
        if not all([self.bd_panjang.text(), self.bd_tinggi.text()]):
            raise ValueError("Panjang dan tinggi harus diisi!")
        
        panjang = float(self.bd_panjang.text())
        tinggi = float(self.bd_tinggi.text())
        bd_type = self.bd_type.currentText()
        
        area = (panjang / 100) * (tinggi / 100)
        
        if bd_type == 'PVC':
            unit_price = 850000
            finishing = 'PVC'
        else:
            finishing = self.bd_finishing.currentText()
            if bd_type == 'Flat':
                prices = {'HPL': 1800000, 'Duco': 2500000, 'Kombinasi': 2100000}
            else:  # Tebal
                prices = {'HPL': 2500000, 'Duco': 3600000, 'Kombinasi': 3200000}
            unit_price = prices.get(finishing, 0)
        
        total = area * unit_price
        
        self.list_widget.addItem(
            f"Backdrop Panel - {bd_type} | {finishing} | {panjang}x{tinggi}cm | Rp{total:,.0f}"
        )
        
        self.items_data.append({
            'nama_item': 'Backdrop Panel',
            'sub_item': bd_type,
            'deskripsi': MaterialDescriptions.get_description(finishing),
            'dimensi': f"{panjang} x {tinggi}",
            'satuan_dimensi': 'cm',
            'total_volume': area,
            'harga_dasar': unit_price,
            'jumlah': total
        })
        
        self.bd_panjang.clear()
        self.bd_tinggi.clear()

    def add_credenza_item(self):
        if not all([self.cr_panjang.text(), self.cr_tinggi.text(), self.cr_tebal.text()]):
            raise ValueError("Semua field harus diisi!")
        
        panjang = float(self.cr_panjang.text())
        tinggi = float(self.cr_tinggi.text())
        tebal = float(self.cr_tebal.text())
        finishing = self.cr_finishing.currentText()
        
        base_total, area, base_unit = self.calculate_cabinet_price(panjang, tinggi, tebal, finishing)
        
        # Credenza: 25% cheaper
        total = base_total * 0.75
        unit_price = base_unit * 0.75
        
        self.list_widget.addItem(
            f"Credenza | {finishing} | {max(panjang,100)}x{max(tinggi,100)}x{tebal}cm | Rp{total:,.0f}"
        )
        
        self.items_data.append({
            'nama_item': 'Credenza',
            'sub_item': '-',
            'deskripsi': MaterialDescriptions.get_description(finishing),
            'dimensi': f"{max(panjang,100)} x {max(tinggi,100)} x {tebal}",
            'satuan_dimensi': 'cm',
            'total_volume': area,
            'harga_dasar': unit_price,
            'jumlah': total
        })
        
        self.cr_panjang.clear()
        self.cr_tinggi.clear()
        self.cr_tebal.clear()

    def add_multi_cabinet_item(self):
        if not all([self.mc_panjang.text(), self.mc_tinggi.text(), self.mc_tebal.text()]):
            raise ValueError("Semua field harus diisi!")
        
        panjang = float(self.mc_panjang.text())
        tinggi = float(self.mc_tinggi.text())
        tebal = float(self.mc_tebal.text())
        finishing = self.mc_finishing.currentText()
        
        base_total, area, base_unit = self.calculate_cabinet_price(panjang, tinggi, tebal, finishing)
        
        # Multi Cabinet: 15% cheaper
        total = base_total * 0.85
        unit_price = base_unit * 0.85
        
        self.list_widget.addItem(
            f"Multi Cabinet | {finishing} | {max(panjang,100)}x{max(tinggi,100)}x{tebal}cm | Rp{total:,.0f}"
        )
        
        self.items_data.append({
            'nama_item': 'Multi Cabinet',
            'sub_item': '-',
            'deskripsi': MaterialDescriptions.get_description(finishing),
            'dimensi': f"{max(panjang,100)} x {max(tinggi,100)} x {tebal}",
            'satuan_dimensi': 'cm',
            'total_volume': area,
            'harga_dasar': unit_price,
            'jumlah': total
        })
        
        self.mc_panjang.clear()
        self.mc_tinggi.clear()
        self.mc_tebal.clear()

    def add_custom_item(self):
        if not all([self.custom_nama.text(), self.custom_panjang.text(), self.custom_unit_price.text()]):
            raise ValueError("Nama, dimensi, dan harga satuan harus diisi!")
        
        nama = self.custom_nama.text()
        panjang = float(self.custom_panjang.text())
        lebar = float(self.custom_lebar.text()) if self.custom_lebar.text() else 0
        tinggi = float(self.custom_tinggi.text()) if self.custom_tinggi.text() else 0
        method = self.custom_method.currentText()
        unit_price = float(self.custom_unit_price.text())
        
        if method == 'Meter Lari':
            volume = panjang / 100
            dimensi = f"{panjang}"
            satuan = "cm"
        elif method == 'Meter Persegi':
            panjang = max(panjang, 100)
            lebar = max(lebar, 100)
            volume = (panjang / 100) * (lebar / 100)
            dimensi = f"{panjang} x {lebar}"
            satuan = "cm"
        elif method == 'Side Area X':
            volume = (lebar / 100) * (tinggi / 100)
            dimensi = f"{panjang} x {lebar} x {tinggi}"
            satuan = "cm"
        elif method == 'Side Area Y':
            volume = (panjang / 100) * (tinggi / 100)
            dimensi = f"{panjang} x {lebar} x {tinggi}"
            satuan = "cm"
        else:  # Side Area Z
            volume = (panjang / 100) * (lebar / 100)
            dimensi = f"{panjang} x {lebar} x {tinggi}"
            satuan = "cm"
        
        total = volume * unit_price
        
        self.list_widget.addItem(
            f"Custom - {nama} | {method} | {dimensi}{satuan} | Rp{total:,.0f}"
        )
        
        self.items_data.append({
            'nama_item': 'Custom Furniture',
            'sub_item': nama,
            'deskripsi': f"Metode: {method}",
            'dimensi': dimensi,
            'satuan_dimensi': satuan,
            'total_volume': volume,
            'harga_dasar': unit_price,
            'jumlah': total
        })
        
        self.custom_nama.clear()
        self.custom_panjang.clear()
        self.custom_lebar.clear()
        self.custom_tinggi.clear()
        self.custom_unit_price.clear()

    def generate_total(self):
        if not self.items_data:
            QMessageBox.information(self, 'Total', 'Tidak ada item untuk dihitung!')
            return
        
        subtotal = sum(item['jumlah'] for item in self.items_data)
        discount_amount = subtotal * (self.discount_percent / 100)
        after_discount = subtotal - discount_amount
        
        msg = f"📊 RINCIAN TOTAL:\n\n"
        msg += f"Subtotal: Rp {subtotal:,.0f}\n"
        if self.discount_percent > 0:
            msg += f"Diskon ({self.discount_percent}%): -Rp {discount_amount:,.0f}\n"
            msg += f"Setelah Diskon: Rp {after_discount:,.0f}\n"
        
        if self.use_tax:
            tax_amount = after_discount * (self.tax_percent / 100)
            grand_total = after_discount + tax_amount
            msg += f"Pajak/PPN ({self.tax_percent}%): +Rp {tax_amount:,.0f}\n"
        else:
            grand_total = after_discount
            msg += f"Pajak/PPN: Tidak Digunakan\n"
        
        msg += f"\n{'='*40}\n"
        msg += f"GRAND TOTAL: Rp {grand_total:,.0f}\n"
        msg += f"{'='*40}\n\n"
        msg += f"Total Item: {len(self.items_data)}"
        
        QMessageBox.information(self, 'Total Harga', msg)

    def export_to_excel(self):
        if not self.items_data:
            QMessageBox.information(self, 'Export ke Excel', 
                                  'Tidak ada data untuk diexport. Tambahkan item terlebih dahulu.')
            return

        # Generate filename: MPN(tahun)(bulan romawi)(no urut)-(nama customer)
        from datetime import datetime
        now = datetime.now()
        year = now.strftime('%Y')
        
        # Bulan dalam angka romawi
        roman_months = ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII']
        month_roman = roman_months[now.month - 1]
        
        # No urut (untuk sementara gunakan hari + jam menit untuk unique)
        sequence = now.strftime('%d%H%M')
        
        # Nama customer (atau "NoName" jika kosong)
        customer_name = self.customer_name.strip() if self.customer_name else "NoName"
        # Bersihkan karakter tidak valid untuk filename
        customer_name = "".join(c for c in customer_name if c.isalnum() or c in (' ', '-', '_')).strip()
        
        default_name = f'MPN{year}{month_roman}{sequence}-{customer_name}.xlsx'
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Simpan File Excel', default_name, 'Excel Files (*.xlsx)'
        )
        
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            from datetime import datetime
        except ImportError:
            QMessageBox.critical(
                self, 'openpyxl belum terpasang',
                'Library openpyxl belum terpasang. Silakan install dengan:\npython -m pip install openpyxl'
            )
            return

        wb = Workbook()
        ws = wb.active
        ws.title = 'Daftar Harga'

        # ========== LOGO & COMPANY HEADER ==========
        # Tambahkan logo jika ada
        logo_path = os.path.join(os.path.dirname(__file__), 'logo.png')
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                # Set ukuran logo H=2.7cm (~102 pixels), W=7cm (~264 pixels)
                # 1 cm ≈ 37.8 pixels di Excel
                img.width = 264  # 7cm
                img.height = 102  # 2.7cm
                # Posisi logo di A1
                ws.add_image(img, 'A1')
                # Set tinggi row untuk logo
                ws.row_dimensions[1].height = 60
                ws.row_dimensions[2].height = 20
            except Exception as e:
                print(f"Gagal menambahkan logo: {e}")
        
        # ========== COMPANY INFO (rata kanan, sejajar dengan logo) ==========
        # Merge cells untuk header perusahaan (mulai dari kolom D sampai I)
        ws.merge_cells('D1:I1')
        ws.merge_cells('D2:I2')
        ws.merge_cells('D3:I3')
        ws.merge_cells('D4:I4')
        
        # Company name (rata kanan)
        ws['D1'] = 'PT. ENIGMA PRISMA DELAPAN'
        ws['D1'].font = Font(bold=True, size=14, color='c0392b')
        ws['D1'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Address line 1 (rata kanan)
        ws['D2'] = 'Jl. Raya H. Abdullah No.56, Pakulonan Barat, Kelapa Dua'
        ws['D2'].font = Font(size=9)
        ws['D2'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Address line 2 (rata kanan)
        ws['D3'] = 'Tangerang, Banten 15812 - 0821 1213 4258'
        ws['D3'].font = Font(size=9)
        ws['D3'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Website & social media (rata kanan)
        ws['D4'] = 'interiormapan.com - mapan.interiorr'
        ws['D4'].font = Font(size=9, color='e74c3c', bold=True)
        ws['D4'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Add spacing
        ws.row_dimensions[5].height = 5
        
        # Date
        ws.merge_cells('A6:I6')
        ws['A6'] = f'Tanggal: {datetime.now().strftime("%d %B %Y, %H:%M")}'
        ws['A6'].font = Font(size=10, italic=True)
        ws['A6'].alignment = Alignment(horizontal='right')
        
        # Add spacing
        ws.row_dimensions[7].height = 10
        
        # ========== TABLE HEADERS ==========
        headers = ['No.', 'Nama Item', 'Sub Item', 'Deskripsi', 'Dimensi', 
                  'Satuan Dimensi', 'Total Volume', 'Harga Dasar', 'Jumlah']
        
        header_row = 8
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # ========== DATA ROWS ==========
        for idx, item in enumerate(self.items_data, start=1):
            row = [
                idx,
                item['nama_item'],
                item['sub_item'],
                item['deskripsi'],
                item['dimensi'],
                item['satuan_dimensi'],
                float(item['total_volume']),
                float(item['harga_dasar']),
                float(item['jumlah'])
            ]
            
            row_num = header_row + idx
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Harga (H dan I) rata kanan, sisanya center
                if col_idx in [8, 9]:  # Harga Dasar dan Jumlah
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Number formatting dengan format Rupiah
        for r in range(header_row + 1, header_row + 1 + len(self.items_data)):
            ws[f'G{r}'].number_format = '0.00'
            ws[f'H{r}'].number_format = '"Rp "#,##0'
            ws[f'I{r}'].number_format = '"Rp "#,##0'
        
        # ========== TOTAL ROWS ==========
        total_row = header_row + len(self.items_data) + 2
        
        subtotal = sum(item['jumlah'] for item in self.items_data)
        
        # Subtotal (rata kanan)
        ws.merge_cells(f'A{total_row}:H{total_row}')
        ws[f'A{total_row}'] = 'SUBTOTAL'
        ws[f'A{total_row}'].font = Font(bold=True, size=11)
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'I{total_row}'] = subtotal
        ws[f'I{total_row}'].font = Font(bold=True, size=11)
        ws[f'I{total_row}'].number_format = '"Rp "#,##0'
        ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        total_row += 1
        
        # Discount (if any)
        if self.discount_percent > 0:
            discount_amount = subtotal * (self.discount_percent / 100)
            ws.merge_cells(f'A{total_row}:H{total_row}')
            ws[f'A{total_row}'] = f'DISKON ({self.discount_percent}%)'
            ws[f'A{total_row}'].font = Font(size=10, color='e74c3c')
            ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'I{total_row}'] = -discount_amount
            ws[f'I{total_row}'].font = Font(size=10, color='e74c3c')
            ws[f'I{total_row}'].number_format = '"Rp "#,##0'
            ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
            total_row += 1
            
            after_discount = subtotal - discount_amount
        else:
            after_discount = subtotal
        
        # Tax (jika digunakan)
        if self.use_tax:
            tax_amount = after_discount * (self.tax_percent / 100)
            ws.merge_cells(f'A{total_row}:H{total_row}')
            ws[f'A{total_row}'] = f'PAJAK/PPN ({self.tax_percent}%)'
            ws[f'A{total_row}'].font = Font(size=10)
            ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'I{total_row}'] = tax_amount
            ws[f'I{total_row}'].font = Font(size=10)
            ws[f'I{total_row}'].number_format = '"Rp "#,##0'
            ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
            total_row += 1
            grand_total = after_discount + tax_amount
        else:
            grand_total = after_discount
        
        # Grand Total (warna merah)
        ws.merge_cells(f'A{total_row}:H{total_row}')
        ws[f'A{total_row}'] = 'GRAND TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True, size=13, color='FFFFFF')
        ws[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{total_row}'].fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
        ws[f'I{total_row}'] = grand_total
        ws[f'I{total_row}'].font = Font(bold=True, size=13, color='FFFFFF')
        ws[f'I{total_row}'].number_format = '"Rp "#,##0'
        ws[f'I{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'I{total_row}'].fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
        
        # ========== CUSTOMER INFO (setelah grand total, rata kiri) ==========
        if self.customer_name or self.customer_address or self.customer_phone:
            total_row += 2
            ws.merge_cells(f'A{total_row}:I{total_row}')
            ws[f'A{total_row}'] = '👤 INFORMASI CUSTOMER'
            ws[f'A{total_row}'].font = Font(bold=True, size=11, color='2c3e50')
            ws[f'A{total_row}'].fill = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')
            ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            total_row += 1
            if self.customer_name:
                ws.merge_cells(f'A{total_row}:I{total_row}')
                ws[f'A{total_row}'] = f'Nama: {self.customer_name}'
                ws[f'A{total_row}'].font = Font(size=10)
                ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
                total_row += 1
            
            if self.customer_address:
                ws.merge_cells(f'A{total_row}:I{total_row}')
                ws[f'A{total_row}'] = f'Alamat: {self.customer_address}'
                ws[f'A{total_row}'].font = Font(size=10)
                ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
                total_row += 1
            
            if self.customer_phone:
                ws.merge_cells(f'A{total_row}:I{total_row}')
                ws[f'A{total_row}'] = f'Telepon: {self.customer_phone}'
                ws[f'A{total_row}'].font = Font(size=10)
                ws[f'A{total_row}'].alignment = Alignment(horizontal='left', vertical='center')
                total_row += 1

        # ========== COLUMN WIDTHS ==========
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15

        try:
            wb.save(file_path)
            QMessageBox.information(self, 'Export ke Excel', 
                                  f'Berhasil menyimpan ke:\n{file_path}')
        except Exception as e:
            QMessageBox.critical(self, 'Gagal Menyimpan', f'Gagal menyimpan file: {e}')

    def toggle_dark_mode(self):
        """Toggle between dark mode and light mode"""
        self.is_dark_mode = not self.is_dark_mode
        if self.is_dark_mode:
            self.apply_dark_mode()
            self.theme_toggle_btn.setText('☀️')
            self.theme_toggle_btn.setToolTip('Switch to Light Mode')
            self.update_info_labels_dark()
        else:
            self.apply_light_mode()
            self.theme_toggle_btn.setText('🌙')
            self.theme_toggle_btn.setToolTip('Switch to Dark Mode')
            self.update_info_labels_light()

    def apply_light_mode(self):
        """Apply light mode theme"""
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f6fa;
                font-size: 12px;
                color: #2c3e50;
            }
            QLabel#TitleLabel {
                color: #c0392b;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border-radius: 4px;
                padding: 8px 16px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #ff6b6b;
            }
            QPushButton#ThemeToggle {
                background-color: #3498db;
                font-size: 12px;
                padding: 6px 12px;
            }
            QPushButton#ThemeToggle:hover {
                background-color: #5dade2;
            }
            QLineEdit, QComboBox {
                border: 1px solid #dcdde1;
                border-radius: 4px;
                padding: 6px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
            }
            QListWidget {
                border: 1px solid #dcdde1;
                border-radius: 4px;
                font-size: 14px;
                background: #fff;
                color: #2c3e50;
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #dcdde1;
                border-radius: 4px;
                margin-top: 12px;
                color: #2c3e50;
                background-color: transparent;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 7px;
                padding: 0 5px 0 5px;
            }
            QScrollArea {
                border: none;
                background-color: transparent;
            }
        """)

    def apply_dark_mode(self):
        """Apply dark mode theme"""
        self.setStyleSheet("""
            QWidget {
                background-color: #1e1e1e;
                font-size: 12px;
                color: #e0e0e0;
            }
            QLabel#TitleLabel {
                color: #9e9e9e;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton {
                background-color: #616161;
                color: white;
                border-radius: 4px;
                padding: 8px 16px;
                font-size: 14px;
                border: 1px solid #424242;
            }
            QPushButton:hover {
                background-color: #757575;
            }
            QPushButton#ThemeToggle {
                background-color: #9e9e9e;
                font-size: 12px;
                padding: 6px 12px;
                border: 1px solid #757575;
            }
            QPushButton#ThemeToggle:hover {
                background-color: #bdbdbd;
            }
            QLineEdit, QComboBox {
                border: 1px solid #444;
                border-radius: 4px;
                padding: 6px;
                font-size: 14px;
                background-color: #2d2d2d;
                color: #e0e0e0;
            }
            QComboBox::drop-down {
                border: none;
                background-color: #3d3d3d;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #e0e0e0;
                width: 0px;
                height: 0px;
            }
            QComboBox QAbstractItemView {
                background-color: #2d2d2d;
                color: #e0e0e0;
                selection-background-color: #3d3d3d;
                border: 1px solid #444;
            }
            QListWidget {
                border: 1px solid #444;
                border-radius: 4px;
                font-size: 14px;
                background: #2d2d2d;
                color: #e0e0e0;
            }
            QListWidget::item:selected {
                background-color: #3d3d3d;
                color: #bdbdbd;
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #444;
                border-radius: 4px;
                margin-top: 12px;
                color: #e0e0e0;
                background-color: #252525;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 7px;
                padding: 0 5px 0 5px;
                color: #9e9e9e;
            }
            QScrollArea {
                border: none;
                background-color: #1e1e1e;
            }
            QLabel {
                color: #e0e0e0;
            }
        """)

    def update_info_labels_light(self):
        """Update all info labels to light mode styling"""
        # Update company info
        if hasattr(self, 'company_info'):
            self.company_info.setStyleSheet("""
                QLabel {
                    color: #2c3e50;
                    font-size: 9px;
                    background-color: white;
                    border-radius: 4px;
                    padding: 6px;
                }
            """)
            self.company_info.setText(
                "<div style='text-align: right;'>"
                "<b style='font-size: 11px; color: #c0392b;'>PT. ENIGMA PRISMA DELAPAN</b><br>"
                "<span style='font-size: 8px;'>Jl. Raya H. Abdullah No.56, Pakulonan Barat<br>"
                "Tangerang - 0821 1213 4258</span>"
                "</div>"
            )
        
        # Update separator
        if hasattr(self, 'separator'):
            self.separator.setStyleSheet("background-color: #c0392b; max-height: 3px;")
        
        # Update total display
        if hasattr(self, 'total_display'):
            self.total_display.setStyleSheet("""
                QLabel {
                    background-color: #e8f5e9;
                    border: 2px solid #4caf50;
                    border-radius: 6px;
                    padding: 12px;
                    font-size: 16px;
                    font-weight: bold;
                    color: #2e7d32;
                }
            """)
        
        light_style = """
            QLabel {
                background-color: #e8f5e9;
                border: 1px solid #4caf50;
                border-radius: 4px;
                padding: 8px;
                color: #2e7d32;
                font-size: 10px;
            }
        """
        
        # Update HPL info labels
        if hasattr(self, 'kb_hpl_info'):
            self.kb_hpl_info.setStyleSheet(light_style)
        if hasattr(self, 'ka_hpl_info'):
            self.ka_hpl_info.setStyleSheet(light_style)
        if hasattr(self, 'wr_hpl_info'):
            self.wr_hpl_info.setStyleSheet(light_style)
        if hasattr(self, 'cr_hpl_info'):
            self.cr_hpl_info.setStyleSheet(light_style)
        if hasattr(self, 'mc_hpl_info'):
            self.mc_hpl_info.setStyleSheet(light_style)
        
        # Update backdrop type info
        if hasattr(self, 'bd_type_info'):
            self.bd_type_info.setStyleSheet("""
                QLabel {
                    background-color: #e8f4f8;
                    border: 1px solid #4a90e2;
                    border-radius: 4px;
                    padding: 8px;
                    color: #2c5aa0;
                    font-size: 11px;
                }
            """)
        
        # Update custom method info
        if hasattr(self, 'custom_method_info'):
            self.custom_method_info.setStyleSheet("""
                QLabel {
                    background-color: #fff3cd;
                    border: 1px solid #ffc107;
                    border-radius: 4px;
                    padding: 8px;
                    color: #856404;
                    font-size: 11px;
                }
            """)

    def update_info_labels_dark(self):
        """Update all info labels to dark mode styling"""
        # Update company info
        if hasattr(self, 'company_info'):
            self.company_info.setStyleSheet("""
                QLabel {
                    color: #e0e0e0;
                    font-size: 9px;
                    background-color: #2d2d2d;
                    border-radius: 4px;
                    padding: 6px;
                    border: 1px solid #444;
                }
            """)
            self.company_info.setText(
                "<div style='text-align: right;'>"
                "<b style='font-size: 11px; color: #bdbdbd;'>PT. ENIGMA PRISMA DELAPAN</b><br>"
                "<span style='font-size: 8px; color: #e0e0e0;'>Jl. Raya H. Abdullah No.56, Pakulonan Barat<br>"
                "Tangerang - 0821 1213 4258</span>"
                "</div>"
            )
        
        # Update separator
        if hasattr(self, 'separator'):
            self.separator.setStyleSheet("background-color: #616161; max-height: 2px;")
        
        # Update total display
        if hasattr(self, 'total_display'):
            self.total_display.setStyleSheet("""
                QLabel {
                    background-color: #1a3a1a;
                    border: 2px solid #4caf50;
                    border-radius: 6px;
                    padding: 12px;
                    font-size: 16px;
                    font-weight: bold;
                    color: #a5d6a7;
                }
            """)
        
        dark_style = """
            QLabel {
                background-color: #1a3a1a;
                border: 1px solid #4caf50;
                border-radius: 4px;
                padding: 8px;
                color: #a5d6a7;
                font-size: 10px;
            }
        """
        
        # Update HPL info labels
        if hasattr(self, 'kb_hpl_info'):
            self.kb_hpl_info.setStyleSheet(dark_style)
        if hasattr(self, 'ka_hpl_info'):
            self.ka_hpl_info.setStyleSheet(dark_style)
        if hasattr(self, 'wr_hpl_info'):
            self.wr_hpl_info.setStyleSheet(dark_style)
        if hasattr(self, 'cr_hpl_info'):
            self.cr_hpl_info.setStyleSheet(dark_style)
        if hasattr(self, 'mc_hpl_info'):
            self.mc_hpl_info.setStyleSheet(dark_style)
        
        # Update backdrop type info
        if hasattr(self, 'bd_type_info'):
            self.bd_type_info.setStyleSheet("""
                QLabel {
                    background-color: #1a2838;
                    border: 1px solid #4a90e2;
                    border-radius: 4px;
                    padding: 8px;
                    color: #90caf9;
                    font-size: 11px;
                }
            """)
        
        # Update custom method info
        if hasattr(self, 'custom_method_info'):
            self.custom_method_info.setStyleSheet("""
                QLabel {
                    background-color: #3a3020;
                    border: 1px solid #ffc107;
                    border-radius: 4px;
                    padding: 8px;
                    color: #ffca28;
                    font-size: 11px;
                }
            """)

    def setup_shortcuts(self):
        """Setup keyboard shortcuts"""
        # Delete selected item
        delete_shortcut = QShortcut(QKeySequence.Delete, self)
        delete_shortcut.activated.connect(self.delete_selected_item)
        
        # Save project
        save_shortcut = QShortcut(QKeySequence.Save, self)
        save_shortcut.activated.connect(self.save_project)
        
        # Open project
        open_shortcut = QShortcut(QKeySequence.Open, self)
        open_shortcut.activated.connect(self.load_project)
        
        # Export
        export_shortcut = QShortcut(QKeySequence("Ctrl+E"), self)
        export_shortcut.activated.connect(self.export_to_excel)

    def update_total_display(self):
        """Update the running total display"""
        if not self.items_data:
            self.total_display.setText("📊 Total: Rp 0 (0 item)")
            return
        
        subtotal = sum(item['jumlah'] for item in self.items_data)
        discount_amount = subtotal * (self.discount_percent / 100)
        after_discount = subtotal - discount_amount
        
        display_text = f"📊 Subtotal: Rp {subtotal:,.0f} | "
        if self.discount_percent > 0:
            display_text += f"Diskon ({self.discount_percent}%): -Rp {discount_amount:,.0f} | "
        
        if self.use_tax:
            tax_amount = after_discount * (self.tax_percent / 100)
            grand_total = after_discount + tax_amount
            display_text += f"Pajak ({self.tax_percent}%): +Rp {tax_amount:,.0f} | "
        else:
            grand_total = after_discount
            display_text += "Pajak: Tidak Digunakan | "
        
        display_text += f"<b>TOTAL: Rp {grand_total:,.0f}</b> ({len(self.items_data)} item)"
        
        self.total_display.setText(display_text)

    def delete_selected_item(self):
        """Delete the selected item from the list"""
        current_row = self.list_widget.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, 'Peringatan', 'Pilih item yang ingin dihapus!')
            return
        
        reply = QMessageBox.question(self, 'Konfirmasi', 
                                     'Hapus item yang dipilih?',
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.list_widget.takeItem(current_row)
            self.items_data.pop(current_row)
            self.update_total_display()
            QMessageBox.information(self, 'Sukses', 'Item berhasil dihapus!')

    def clear_all_items(self):
        """Clear all items from the list"""
        if not self.items_data:
            QMessageBox.information(self, 'Info', 'Tidak ada item untuk dihapus!')
            return
        
        reply = QMessageBox.question(self, 'Konfirmasi', 
                                     f'Hapus semua {len(self.items_data)} item?',
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.list_widget.clear()
            self.items_data.clear()
            self.update_total_display()
            QMessageBox.information(self, 'Sukses', 'Semua item berhasil dihapus!')

    def show_customer_dialog(self):
        """Show dialog for customer information"""
        dialog = QDialog(self)
        dialog.setWindowTitle('Informasi Customer')
        dialog.setMinimumWidth(400)
        
        layout = QFormLayout()
        
        name_edit = QLineEdit(self.customer_name)
        name_edit.setPlaceholderText("Nama customer")
        layout.addRow('Nama:', name_edit)
        
        address_edit = QLineEdit(self.customer_address)
        address_edit.setPlaceholderText("Alamat lengkap")
        layout.addRow('Alamat:', address_edit)
        
        phone_edit = QLineEdit(self.customer_phone)
        phone_edit.setPlaceholderText("Nomor telepon")
        layout.addRow('Telepon:', phone_edit)
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            self.customer_name = name_edit.text()
            self.customer_address = address_edit.text()
            self.customer_phone = phone_edit.text()
            QMessageBox.information(self, 'Sukses', 'Info customer berhasil disimpan!')

    def show_discount_dialog(self):
        """Show dialog for discount and tax settings"""
        from PyQt5.QtWidgets import QCheckBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle('Pengaturan Diskon & Pajak')
        dialog.setMinimumWidth(350)
        
        layout = QFormLayout()
        
        discount_spin = QDoubleSpinBox()
        discount_spin.setRange(0, 100)
        discount_spin.setSuffix(' %')
        discount_spin.setValue(self.discount_percent)
        discount_spin.setDecimals(2)
        layout.addRow('Diskon:', discount_spin)
        
        # Checkbox untuk aktifkan pajak
        tax_checkbox = QCheckBox('Gunakan Pajak/PPN')
        tax_checkbox.setChecked(self.use_tax)
        layout.addRow(tax_checkbox)
        
        tax_spin = QDoubleSpinBox()
        tax_spin.setRange(0, 100)
        tax_spin.setSuffix(' %')
        tax_spin.setValue(self.tax_percent)
        tax_spin.setDecimals(2)
        tax_spin.setEnabled(self.use_tax)
        layout.addRow('Pajak (PPN):', tax_spin)
        
        # Toggle tax input enabled/disabled
        def toggle_tax():
            tax_spin.setEnabled(tax_checkbox.isChecked())
        
        tax_checkbox.stateChanged.connect(toggle_tax)
        
        info_label = QLabel('💡 Diskon akan dikurangi dari subtotal,\nlalu pajak dihitung dari hasil setelah diskon.')
        info_label.setStyleSheet("color: #666; font-size: 10px;")
        layout.addRow(info_label)
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            self.discount_percent = discount_spin.value()
            self.tax_percent = tax_spin.value()
            self.use_tax = tax_checkbox.isChecked()
            self.update_total_display()
            
            tax_status = f"Pajak: {self.tax_percent}%" if self.use_tax else "Pajak: Tidak Digunakan"
            QMessageBox.information(self, 'Sukses', 
                f'Diskon: {self.discount_percent}% | {tax_status}')

    def save_project(self):
        """Save current project to JSON file"""
        if not self.items_data:
            QMessageBox.warning(self, 'Peringatan', 'Tidak ada data untuk disimpan!')
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Save Project', '', 'JSON Files (*.json)'
        )
        
        if file_path:
            try:
                project_data = {
                    'customer_name': self.customer_name,
                    'customer_address': self.customer_address,
                    'customer_phone': self.customer_phone,
                    'discount_percent': self.discount_percent,
                    'tax_percent': self.tax_percent,
                    'use_tax': self.use_tax,
                    'items': self.items_data,
                    'saved_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(project_data, f, indent=2, ensure_ascii=False)
                
                QMessageBox.information(self, 'Sukses', 
                    f'Project berhasil disimpan!\n{file_path}')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Gagal menyimpan: {e}')

    def load_project(self):
        """Load project from JSON file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, 'Load Project', '', 'JSON Files (*.json)'
        )
        
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    project_data = json.load(f)
                
                # Clear current data
                self.list_widget.clear()
                self.items_data.clear()
                
                # Load customer info
                self.customer_name = project_data.get('customer_name', '')
                self.customer_address = project_data.get('customer_address', '')
                self.customer_phone = project_data.get('customer_phone', '')
                self.discount_percent = project_data.get('discount_percent', 0.0)
                self.tax_percent = project_data.get('tax_percent', 11.0)
                self.use_tax = project_data.get('use_tax', True)
                
                # Load items
                self.items_data = project_data.get('items', [])
                
                # Populate list widget
                for item in self.items_data:
                    display_text = f"{item['nama_item']} - {item['sub_item']} | {item['dimensi']}{item['satuan_dimensi']} | Rp{item['jumlah']:,.0f}"
                    self.list_widget.addItem(display_text)
                
                self.update_total_display()
                
                saved_at = project_data.get('saved_at', 'Unknown')
                QMessageBox.information(self, 'Sukses', 
                    f'Project berhasil dimuat!\nDisimpan: {saved_at}\nTotal: {len(self.items_data)} item')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Gagal memuat: {e}')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = FurniturePriceGenerator()
    window.show()
    sys.exit(app.exec_())
